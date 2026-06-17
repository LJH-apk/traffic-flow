"""
人工复核集验证工具。

该脚本用于验证已有算法输出，而不是替代主办方隐藏评分。核心目标是把
cross_section.csv 与人工过线标注对齐，输出可写入报告的汇总指标，以及
后续绘图可直接使用的明细 CSV/JSON。

人工过线标注 CSV 至少包含：
    gt_frame_id, section, class_name, direction

建议额外包含 track_id；若缺失，匹配会退化为 section + frame window。
"""
from __future__ import annotations

import argparse
import csv
import json
import math
from dataclasses import dataclass
from pathlib import Path
from statistics import mean
from typing import Iterable


DEFAULT_OUTPUT_DIR = Path("outputs") / "validation"
DEFAULT_NORMALIZED_MANUAL_CSV = Path("annotations") / "manual_crossing.csv"
_ALGO_SECTION_SET = {"北进口主断面", "北进口右转", "北进口掉头车道"}
_MAIN_SECTION = "北进口主断面"
_RIGHT_SECTION = "北进口右转"
_DEDUP_GAP_FRAMES = 25
_MATCH_CLASS_PENALTY = 1000
_MATCH_DIRECTION_PENALTY = 100
_MATCH_LANE_PENALTY = 20
_MANUAL_FIELDS = [
    "entrance", "video_name", "gt_frame_id", "section", "direction",
    "class_name", "lane_id", "track_id", "vehicle_label", "note",
]


@dataclass(frozen=True)
class EventMatch:
    match_status: str
    pred_index: int | None
    gt_index: int | None
    section: str
    track_id: str
    lane_id: str          # pred_lane_id（优先），退化到 gt_lane_id
    class_name: str        # pred_class_name
    direction: str         # pred_direction
    pred_frame_id: int | None
    gt_frame_id: int | None
    frame_error_s: float | None
    lane_correct: int | None
    direction_correct: int | None
    class_correct: int | None
    # 以下为扩展诊断字段
    gt_lane_id: str = ""
    gt_class_name: str = ""
    gt_direction: str = ""
    pred_lane_id: str = ""
    pred_class_name: str = ""
    pred_direction: str = ""


def _read_csv(path: Path | None) -> list[dict[str, str]]:
    if path is None or not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def _write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _to_int(value: object, default: int | None = None) -> int | None:
    try:
        if value in (None, ""):
            return default
        return int(float(str(value)))
    except (TypeError, ValueError):
        return default


def _to_float(value: object, default: float | None = None) -> float | None:
    try:
        if value in (None, ""):
            return default
        return float(str(value))
    except (TypeError, ValueError):
        return default


def _fmt(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return round(value, 4)
    return value


def _normalize_section(value: object) -> str:
    raw = str(value or "").strip()
    if raw in {"主", "主断面", "北进口主断面", "主路", "主线", "section_name"}:
        return _MAIN_SECTION
    if raw in {"右", "右转", "右转道", "北进口右转", "北进口右转专用道"}:
        return _RIGHT_SECTION
    if raw in {"掉头", "掉头车道", "北进口掉头车道"}:
        return "北进口掉头车道"
    return raw


def _normalize_class(row: dict[str, str]) -> str:
    raw = str(row.get("class_name", "") or "").strip().lower()
    label = str(row.get("vehicle_label", row.get("vehicle_name", "")) or "").strip()

    if raw in {"class_name", "vehicle_class"}:
        return "unknown"
    if raw in {"car", "truck", "bus", "motorcycle", "bicycle", "non_motor"}:
        return raw
    if label in {"小汽车", "suv", "轿车", "面包车"}:
        return "car"
    if label in {"货车", "货运皮卡", "卡车"}:
        return "truck"
    if label in {"公交车", "大巴"}:
        return "bus"
    if label in {"自行车"}:
        return "non_motor"
    if label in {"电动车", "三轮车", "摩托车"}:
        return "non_motor"
    return raw or "unknown"


def _normalize_pred_class(row: dict[str, str]) -> str:
    class_name = str(row.get("class_name", "") or "").strip().lower()
    vehicle_category = str(row.get("vehicle_category", "") or "").strip()
    if class_name in {"car", "truck", "bus"}:
        return class_name
    if vehicle_category == "非机动车" or class_name in {"motorcycle", "bicycle"}:
        return "non_motor"
    return class_name or "unknown"


def _vehicle_group(class_name: str) -> str:
    class_name = str(class_name or "").strip().lower()
    if class_name in {"car", "truck", "bus"}:
        return "机动车"
    if class_name in {"motorcycle", "bicycle", "non_motor"}:
        return "非机动车"
    return "其他"


def _class_display_name(class_name: str) -> str:
    class_name = str(class_name or "").strip().lower()
    if class_name in {"motorcycle", "bicycle", "non_motor"}:
        return "non_motor"
    return class_name


def _normalize_lane_id(value: object, section: str = "") -> str:
    raw = str(value or "").strip().upper()
    if raw in {"", "UNKNOWN", "UN", "NAN", "NONE"}:
        return ""
    if raw == "OPPOSITE":
        return "OPPOSITE"
    if raw.startswith("L") and raw[1:].isdigit():
        raw = raw[1:]
    if raw.isdigit():
        lane = raw
    else:
        lane = raw

    # 北进口主断面按人工补充口径：
    # L1 外侧是 OPPOSITE，L4-L5 之间为最外侧非机动车道。
    # 因此人工表里出现的 L5 归并到该非机动车道（车道 4）。
    if section == "北进口主断面" and lane == "5":
        return "4"
    return lane


def _load_table(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - openpyxl should be present
            raise RuntimeError("读取 Excel 需要安装 openpyxl") from exc

        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows: list[dict[str, str]] = []
        for values in ws.iter_rows(values_only=True):
            if not any(v not in (None, "") for v in values):
                continue
            padded = list(values) + [None] * max(0, 10 - len(values))
            rows.append({
                "entrance": str(padded[0] or "").strip(),
                "video_name": str(padded[1] or "").strip(),
                "gt_frame_id": padded[2],
                "section": str(padded[3] or "").strip(),
                "direction": str(padded[4] or "").strip(),
                "class_name": str(padded[5] or "").strip(),
                "lane_id": str(padded[6] or "").strip(),
                "track_id": str(padded[7] or "").strip(),
                "vehicle_label": str(padded[8] or "").strip(),
                "note": str(padded[9] or "").strip(),
            })
        return rows

    if not path.exists():
        return []
    last_error: Exception | None = None
    for encoding in ("utf-8", "utf-8-sig", "gbk", "gb18030"):
        try:
            with path.open(newline="", encoding=encoding) as fh:
                return list(csv.DictReader(fh))
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    if last_error is not None:
        raise last_error
    return []


def export_normalized_manual_annotations(
    manual_crossing_path: Path,
    output_csv: Path = DEFAULT_NORMALIZED_MANUAL_CSV,
) -> list[dict[str, str]]:
    """把人工标注 Excel/CSV 规范化为仓库内 UTF-8 CSV。"""
    raw_events = _load_table(manual_crossing_path)
    events = _normalize_manual_events(raw_events)
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    _write_csv(output_csv, events, _MANUAL_FIELDS)
    return events


def _dedupe_pred_events(pred_events: list[dict[str, str]], gap_frames: int = _DEDUP_GAP_FRAMES) -> list[dict[str, str]]:
    """折叠同一 track+section 在短时间内的重复触发。"""
    if not pred_events:
        return []

    ordered = sorted(
        pred_events,
        key=lambda r: (
            _normalize_section(r.get("section", "")),
            str(r.get("track_id", "")).strip(),
            _to_int(r.get("frame_id"), 0) or 0,
        ),
    )
    collapsed: list[dict[str, str]] = []
    last_by_key: dict[tuple[str, str], tuple[int, dict[str, str]]] = {}
    for event in ordered:
        section = _normalize_section(event.get("section", ""))
        track_id = str(event.get("track_id", "")).strip()
        frame_id = _to_int(event.get("frame_id"))
        if frame_id is None:
            continue
        key = (section, track_id)
        previous = last_by_key.get(key)
        if previous is not None and frame_id - previous[0] <= gap_frames:
            continue
        normalized = dict(event)
        normalized["section"] = section
        last_by_key[key] = (frame_id, normalized)
        collapsed.append(normalized)
    return collapsed


def _normalize_manual_events(raw_events: list[dict[str, str]], allowed_sections: set[str] | None = None) -> list[dict[str, str]]:
    events: list[dict[str, str]] = []
    for row in raw_events:
        entrance = str(row.get("entrance", row.get("video_name", "")) or "").strip()
        if entrance and entrance not in {"北进口", "north", "N"}:
            continue
        frame_id = _to_int(row.get("gt_frame_id", row.get("frame_id")))
        if frame_id is None:
            continue
        raw_section = row.get("section", row.get("section_name", ""))
        section = _normalize_section(raw_section)
        if allowed_sections is not None and section not in allowed_sections:
            continue
        direction = str(row.get("direction", "") or "").strip()
        class_name = _normalize_class(row)
        if direction in {"", "direction"} or class_name == "unknown":
            continue
        events.append({
            "entrance": entrance or "北进口",
            "video_name": str(row.get("video_name", "") or "").strip(),
            "gt_frame_id": str(frame_id),
            "section": section,
            "direction": direction,
            "class_name": class_name,
            "track_id": str(row.get("track_id", row.get("manual_vehicle_id", "")) or "").strip(),
            "vehicle_label": str(row.get("vehicle_label", row.get("note", "")) or "").strip(),
            "lane_id": _normalize_lane_id(row.get("lane_id", ""), section),
            "note": str(row.get("note", "") or "").strip(),
        })
    return events


def _normalize_pred_events(raw_events: list[dict[str, str]]) -> list[dict[str, str]]:
    events: list[dict[str, str]] = []
    for row in raw_events:
        frame_id = _to_int(row.get("frame_id"))
        if frame_id is None:
            continue
        section = _normalize_section(row.get("section", ""))
        if section not in _ALGO_SECTION_SET:
            continue
        events.append({
            "frame_id": str(frame_id),
            "timestamp_s": str(_to_float(row.get("timestamp_s"), 0.0) or 0.0),
            "section": section,
            "track_id": str(row.get("track_id", "") or "").strip(),
            "class_name": _normalize_pred_class(row),
            "lane_id": _normalize_lane_id(row.get("lane_id", ""), section),
            "arrival_departure": str(row.get("arrival_departure", "") or "").strip(),
            "color": str(row.get("color", "") or "").strip(),
            "direction": str(row.get("direction", "") or "").strip(),
            "speed_kmh": str(row.get("speed_kmh", "") or "").strip(),
            "headway_s": str(row.get("headway_s", "") or "").strip(),
            "spacing_m": str(row.get("spacing_m", "") or "").strip(),
        })
    return events


def _direction_for_comparison(section: str, pred: dict[str, str]) -> str:
    """按断面语义选择真正用于比较的方向字段。"""
    if section == _RIGHT_SECTION:
        return str(pred.get("arrival_departure", "")).strip()
    return str(pred.get("direction", "")).strip()


def _pair_cost(pred: dict[str, str], gt: dict[str, str], frame_tolerance: int) -> int | None:
    section = str(gt.get("section", "")).strip()
    pred_frame = _to_int(pred.get("frame_id"))
    gt_frame = _to_int(gt.get("gt_frame_id", gt.get("frame_id")))
    if pred_frame is None or gt_frame is None:
        return None
    err = abs(pred_frame - gt_frame)
    if err > frame_tolerance:
        return None

    gt_class = str(gt.get("class_name", "")).strip()
    pred_class = str(pred.get("class_name", "")).strip()
    gt_lane = _normalize_lane_id(gt.get("lane_id", ""), str(gt.get("section", "")).strip())
    pred_lane = _normalize_lane_id(pred.get("lane_id", ""), str(pred.get("section", "")).strip())
    gt_direction = str(gt.get("direction", "")).strip()
    pred_direction = _direction_for_comparison(section, pred)

    cost = err
    if gt_class and pred_class and gt_class != pred_class:
        cost += _MATCH_CLASS_PENALTY
    if gt_lane and pred_lane and gt_lane != pred_lane:
        cost += _MATCH_LANE_PENALTY
    if gt_direction and pred_direction and gt_direction != pred_direction:
        cost += _MATCH_DIRECTION_PENALTY
    return cost


def _match_events(
    pred_events: list[dict[str, str]],
    gt_events: list[dict[str, str]],
    fps: float,
    frame_tolerance: int,
) -> list[EventMatch]:
    try:
        import numpy as np
        from scipy.optimize import linear_sum_assignment
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("事件最优匹配需要 scipy") from exc

    matches: list[EventMatch] = []
    matched_pred: set[int] = set()
    matched_gt: set[int] = set()
    sections = sorted({str(r.get("section", "")).strip() for r in gt_events + pred_events if str(r.get("section", "")).strip()})

    for section in sections:
        gt_bucket = [(idx, gt) for idx, gt in enumerate(gt_events) if str(gt.get("section", "")).strip() == section]
        pred_bucket = [(idx, pred) for idx, pred in enumerate(pred_events) if str(pred.get("section", "")).strip() == section]
        if not gt_bucket and not pred_bucket:
            continue
        if not gt_bucket:
            continue
        if not pred_bucket:
            continue

        cost = np.full((len(gt_bucket), len(pred_bucket)), 10**9, dtype=int)
        for i, (_, gt) in enumerate(gt_bucket):
            for j, (_, pred) in enumerate(pred_bucket):
                pair_cost = _pair_cost(pred, gt, frame_tolerance)
                if pair_cost is not None:
                    cost[i, j] = pair_cost

        row_ind, col_ind = linear_sum_assignment(cost)
        for i, j in zip(row_ind.tolist(), col_ind.tolist()):
            if cost[i, j] >= 10**9:
                continue
            gt_idx, gt = gt_bucket[i]
            pred_idx, pred = pred_bucket[j]
            matched_gt.add(gt_idx)
            matched_pred.add(pred_idx)
            gt_frame = _to_int(gt.get("gt_frame_id", gt.get("frame_id")))
            pred_frame = _to_int(pred.get("frame_id"))
            gt_direction = str(gt.get("direction", ""))
            pred_direction = _direction_for_comparison(section, pred)
            pred_class = str(pred.get("class_name", ""))
            gt_class = str(gt.get("class_name", ""))
            pred_lane = _normalize_lane_id(pred.get("lane_id", ""), str(pred.get("section", "")).strip())
            gt_lane = _normalize_lane_id(gt.get("lane_id", ""), str(gt.get("section", "")).strip())
            matches.append(EventMatch(
                match_status="TP",
                pred_index=pred_idx,
                gt_index=gt_idx,
                section=section,
                track_id=str(pred.get("track_id", "")),
                lane_id=pred_lane or gt_lane,
                class_name=pred_class,
                direction=pred_direction,
                pred_frame_id=pred_frame,
                gt_frame_id=gt_frame,
                frame_error_s=abs((pred_frame or 0) - (gt_frame or 0)) / fps if pred_frame is not None and gt_frame is not None else None,
                lane_correct=int(pred_lane == gt_lane) if gt_lane else None,
                direction_correct=int(pred_direction == gt_direction) if gt_direction else None,
                class_correct=int(pred_class == gt_class) if gt_class else None,
                gt_lane_id=gt_lane,
                gt_class_name=gt_class,
                gt_direction=gt_direction,
                pred_lane_id=pred_lane,
                pred_class_name=pred_class,
                pred_direction=pred_direction,
            ))

    for gt_idx, gt in enumerate(gt_events):
        if gt_idx in matched_gt:
            continue
        gt_frame = _to_int(gt.get("gt_frame_id", gt.get("frame_id")))
        matches.append(EventMatch(
            match_status="FN",
            pred_index=None,
            gt_index=gt_idx,
            section=str(gt.get("section", "")),
            track_id=str(gt.get("track_id", "")),
            lane_id=_normalize_lane_id(gt.get("lane_id", ""), str(gt.get("section", "")).strip()),
            class_name=str(gt.get("class_name", "")),
            direction=str(gt.get("direction", "")),
            pred_frame_id=None,
            gt_frame_id=gt_frame,
            frame_error_s=None,
            lane_correct=None,
            direction_correct=None,
            class_correct=None,
        ))

    for pred_idx, pred in enumerate(pred_events):
        if pred_idx in matched_pred:
            continue
        section = str(pred.get("section", ""))
        matches.append(EventMatch(
            match_status="FP",
            pred_index=pred_idx,
            gt_index=None,
            section=section,
            track_id=str(pred.get("track_id", "")),
            lane_id=_normalize_lane_id(pred.get("lane_id", ""), str(pred.get("section", "")).strip()),
            class_name=str(pred.get("class_name", "")),
            direction=_direction_for_comparison(section, pred),
            pred_frame_id=_to_int(pred.get("frame_id")),
            gt_frame_id=None,
            frame_error_s=None,
            lane_correct=None,
            direction_correct=None,
            class_correct=None,
        ))

    matches.sort(key=lambda m: ((m.gt_frame_id if m.gt_frame_id is not None else 10**9), (m.pred_frame_id if m.pred_frame_id is not None else 10**9), m.match_status))
    return matches


def _event_match_rows(matches: Iterable[EventMatch]) -> list[dict]:
    rows = []
    for m in matches:
        rows.append({
            "match_status": m.match_status,
            "pred_index": _fmt(m.pred_index),
            "gt_index": _fmt(m.gt_index),
            "section": m.section,
            "track_id": m.track_id,
            "lane_id": m.lane_id,
            "class_name": m.class_name,
            "direction": m.direction,
            "pred_frame_id": _fmt(m.pred_frame_id),
            "gt_frame_id": _fmt(m.gt_frame_id),
            "frame_error_s": _fmt(m.frame_error_s),
            "lane_correct": _fmt(m.lane_correct),
            "direction_correct": _fmt(m.direction_correct),
            "class_correct": _fmt(m.class_correct),
            # 诊断字段：预测 vs 人工标注的原始值
            "pred_lane_id": _fmt(m.pred_lane_id),
            "gt_lane_id": _fmt(m.gt_lane_id),
            "pred_class_name": _fmt(m.pred_class_name),
            "gt_class_name": _fmt(m.gt_class_name),
            "pred_direction": _fmt(m.pred_direction),
            "gt_direction": _fmt(m.gt_direction),
        })
    return rows


def _headway_details(
    gt_events: list[dict[str, str]],
    pred_events: list[dict[str, str]],
    fps: float,
    matches: list[EventMatch] | None = None,
) -> list[dict]:
    pred_by_gt_index: dict[int, dict[str, str]] = {}
    if matches is not None:
        for m in matches:
            if m.match_status != "TP" or m.gt_index is None or m.pred_index is None:
                continue
            if 0 <= m.pred_index < len(pred_events):
                pred_by_gt_index[m.gt_index] = pred_events[m.pred_index]
    sorted_gt = sorted(
        list(enumerate(gt_events)),
        key=lambda r: (
            str(r[1].get("section", "")),
            str(r[1].get("direction", "")),
            _to_int(r[1].get("gt_frame_id", r[1].get("frame_id")), 0) or 0,
        ),
    )
    rows: list[dict] = []
    last_by_key: dict[tuple[str, str], dict[str, str]] = {}
    for orig_idx, gt in sorted_gt:
        section = str(gt.get("section", ""))
        direction = str(gt.get("direction", ""))
        tid = str(gt.get("track_id", ""))
        gt_frame = _to_int(gt.get("gt_frame_id", gt.get("frame_id")))
        if gt_frame is None:
            continue
        key = (section, direction)
        previous = last_by_key.get(key)
        last_by_key[key] = gt
        if previous is None:
            continue
        prev_frame = _to_int(previous.get("gt_frame_id", previous.get("frame_id")))
        if prev_frame is None:
            continue
        gt_headway = (gt_frame - prev_frame) / fps

        pred = pred_by_gt_index.get(orig_idx)
        pred_headway = _to_float(pred.get("headway_s") if pred else None)
        abs_error = abs(pred_headway - gt_headway) if pred_headway is not None else None
        mape = abs_error / gt_headway if abs_error is not None and gt_headway > 0 else None
        rows.append({
            "section": section,
            "direction": direction,
            "track_id": tid,
            "gt_frame_id": gt_frame,
            "previous_gt_frame_id": prev_frame,
            "gt_headway_s": round(gt_headway, 4),
            "pred_headway_s": _fmt(pred_headway),
            "abs_error_s": _fmt(abs_error),
            "mape": _fmt(mape),
            "error_le_0_5s": int(abs_error <= 0.5) if abs_error is not None else "",
        })
    return rows


def _spacing_details(pred_events: list[dict[str, str]], tolerance_m: float = 2.0) -> list[dict]:
    rows: list[dict] = []
    last_by_key: dict[tuple[str, str], dict[str, str]] = {}
    ordered = sorted(
        pred_events,
        key=lambda r: (
            str(r.get("section", "")),
            str(r.get("direction", "")),
            _to_float(r.get("timestamp_s"), 0.0) or 0.0,
            _to_int(r.get("frame_id"), 0) or 0,
        ),
    )
    for event in ordered:
        section = str(event.get("section", ""))
        direction = str(event.get("direction", ""))
        key = (section, direction)
        previous = last_by_key.get(key)
        last_by_key[key] = event
        if previous is None:
            continue
        prev_speed = _to_float(previous.get("speed_kmh"), 0.0) or 0.0
        headway = _to_float(event.get("headway_s"))
        spacing = _to_float(event.get("spacing_m"))
        expected = prev_speed / 3.6 * headway if headway is not None else None
        error = abs(spacing - expected) if spacing is not None and expected is not None else None
        rows.append({
            "section": section,
            "direction": direction,
            "track_id": str(event.get("track_id", "")),
            "frame_id": _fmt(_to_int(event.get("frame_id"))),
            "previous_track_id": str(previous.get("track_id", "")),
            "previous_speed_kmh": _fmt(prev_speed),
            "headway_s": _fmt(headway),
            "spacing_m": _fmt(spacing),
            "expected_spacing_m": _fmt(expected),
            "spacing_error_m": _fmt(error),
            "passed": int(error <= tolerance_m) if error is not None else "",
        })
    return rows


def _anomaly_rows(pred_events: list[dict[str, str]], fps: float) -> list[dict]:
    rows: list[dict] = []
    sorted_events = sorted(
        enumerate(pred_events),
        key=lambda item: (
            str(item[1].get("track_id", "")),
            str(item[1].get("section", "")),
            _to_int(item[1].get("frame_id"), 0) or 0,
        ),
    )
    last_by_track_section: dict[tuple[str, str], tuple[int, dict[str, str]]] = {}

    def add(anomaly_type: str, idx: int, event: dict[str, str], detail: str) -> None:
        rows.append({
            "anomaly_type": anomaly_type,
            "detail": detail,
            "row_index": idx,
            "frame_id": event.get("frame_id", ""),
            "timestamp_s": event.get("timestamp_s", ""),
            "section": event.get("section", ""),
            "track_id": event.get("track_id", ""),
            "class_name": event.get("class_name", ""),
            "direction": event.get("direction", ""),
            "speed_kmh": event.get("speed_kmh", ""),
            "headway_s": event.get("headway_s", ""),
            "spacing_m": event.get("spacing_m", ""),
        })

    for idx, event in sorted_events:
        track_id = str(event.get("track_id", ""))
        section = str(event.get("section", ""))
        frame_id = _to_int(event.get("frame_id"))
        key = (track_id, section)
        previous = last_by_track_section.get(key)
        if previous is not None and frame_id is not None:
            prev_idx, prev_event = previous
            prev_frame = _to_int(prev_event.get("frame_id"))
            if prev_frame is not None and abs(frame_id - prev_frame) <= fps:
                add("duplicate_event", idx, event, f"previous_row_index={prev_idx}")
            if prev_event.get("direction") and event.get("direction") and prev_event.get("direction") != event.get("direction"):
                add("opposite_direction", idx, event, f"previous_row_index={prev_idx}")
        last_by_track_section[key] = (idx, event)

    for idx, event in enumerate(pred_events):
        speed = _to_float(event.get("speed_kmh"))
        headway = _to_float(event.get("headway_s"))
        spacing = _to_float(event.get("spacing_m"))
        if speed is not None and speed > 100:
            add("speed_out_of_range", idx, event, "speed_kmh > 100")
        if headway is not None and 0 < headway < 0.3:
            add("headway_too_short", idx, event, "0 < headway_s < 0.3")
        if spacing is not None and spacing < 0:
            add("negative_spacing", idx, event, "spacing_m < 0")
    return rows


def _summary(matches: list[EventMatch], headway_rows: list[dict], spacing_rows: list[dict], anomalies: list[dict]) -> dict[str, float | int | str]:
    tp = sum(1 for m in matches if m.match_status == "TP")
    fp = sum(1 for m in matches if m.match_status == "FP")
    fn = sum(1 for m in matches if m.match_status == "FN")
    frame_errors = [m.frame_error_s for m in matches if m.match_status == "TP" and m.frame_error_s is not None]
    lane_values = [m.lane_correct for m in matches if m.match_status == "TP" and m.lane_correct is not None]
    direction_values = [m.direction_correct for m in matches if m.match_status == "TP" and m.direction_correct is not None]
    class_values = [m.class_correct for m in matches if m.match_status == "TP" and m.class_correct is not None]
    headway_errors = [_to_float(r.get("abs_error_s")) for r in headway_rows if _to_float(r.get("abs_error_s")) is not None]
    headway_mapes = [_to_float(r.get("mape")) for r in headway_rows if _to_float(r.get("mape")) is not None]
    headway_ok = [_to_int(r.get("error_le_0_5s")) for r in headway_rows if _to_int(r.get("error_le_0_5s")) is not None]
    spacing_pass = [_to_int(r.get("passed")) for r in spacing_rows if _to_int(r.get("passed")) is not None]
    duplicate_count = sum(1 for r in anomalies if r["anomaly_type"] == "duplicate_event")
    physical_count = sum(1 for r in anomalies if r["anomaly_type"] in {"speed_out_of_range", "headway_too_short", "negative_spacing"})

    return {
        "event_tp": tp,
        "event_fp": fp,
        "event_fn": fn,
        "event_precision": tp / (tp + fp) if tp + fp else 0.0,
        "event_recall": tp / (tp + fn) if tp + fn else 0.0,
        "crossing_time_mae_s": mean(frame_errors) if frame_errors else 0.0,
        "lane_accuracy": mean(lane_values) if lane_values else "",
        "direction_accuracy": mean(direction_values) if direction_values else "",
        "class_accuracy": mean(class_values) if class_values else "",
        "headway_mae_s": mean(headway_errors) if headway_errors else 0.0,
        "headway_mape": mean(headway_mapes) if headway_mapes else "",
        "headway_error_le_0_5s_rate": mean(headway_ok) if headway_ok else "",
        "spacing_consistency_pass_rate": mean(spacing_pass) if spacing_pass else "",
        "duplicate_event_count": duplicate_count,
        "physical_anomaly_count": physical_count,
        "total_anomaly_count": len(anomalies),
    }


def _subset_accuracy(values: list[int | None]) -> float | str:
    picked = [v for v in values if v is not None]
    return mean(picked) if picked else ""


def _breakdown_summary(matches: list[EventMatch], gt_events: list[dict[str, str]]) -> dict[str, float | int | str]:
    tp = [m for m in matches if m.match_status == "TP" and m.gt_index is not None]

    def gt_class(m: EventMatch) -> str:
        return str(gt_events[m.gt_index].get("class_name", ""))

    gt_motor = [m for m in tp if gt_class(m) in {"car", "truck", "bus"}]
    gt_car_main = [m for m in tp if gt_class(m) == "car" and m.section == _MAIN_SECTION]

    return {
        "tp_count": len(tp),
        "gt_motor_tp_count": len(gt_motor),
        "gt_motor_class_accuracy": _subset_accuracy([m.class_correct for m in gt_motor]),
        "gt_motor_lane_accuracy": _subset_accuracy([m.lane_correct for m in gt_motor]),
        "gt_motor_direction_accuracy": _subset_accuracy([m.direction_correct for m in gt_motor]),
        "gt_car_main_tp_count": len(gt_car_main),
        "gt_car_main_class_accuracy": _subset_accuracy([m.class_correct for m in gt_car_main]),
        "gt_car_main_lane_accuracy": _subset_accuracy([m.lane_correct for m in gt_car_main]),
        "gt_car_main_direction_accuracy": _subset_accuracy([m.direction_correct for m in gt_car_main]),
    }


def _metric_breakdown_rows(matches: list[EventMatch], gt_events: list[dict[str, str]]) -> list[dict]:
    tp = [m for m in matches if m.match_status == "TP" and m.gt_index is not None]

    def gt_class(m: EventMatch) -> str:
        return str(gt_events[m.gt_index].get("class_name", ""))

    groups = [
        ("overall_tp", tp),
        ("gt_motor_tp", [m for m in tp if gt_class(m) in {"car", "truck", "bus"}]),
        ("gt_nonmotor_tp", [m for m in tp if gt_class(m) not in {"car", "truck", "bus"}]),
        ("gt_car_main_tp", [m for m in tp if gt_class(m) == "car" and m.section == _MAIN_SECTION]),
        ("gt_motor_main_tp", [m for m in tp if gt_class(m) in {"car", "truck", "bus"} and m.section == _MAIN_SECTION]),
    ]
    rows: list[dict] = []
    for scope, subset in groups:
        rows.append({
            "scope": scope,
            "tp_count": len(subset),
            "class_accuracy": _fmt(_subset_accuracy([m.class_correct for m in subset])),
            "lane_accuracy": _fmt(_subset_accuracy([m.lane_correct for m in subset])),
            "direction_accuracy": _fmt(_subset_accuracy([m.direction_correct for m in subset])),
            "crossing_time_mae_s": _fmt(mean([m.frame_error_s for m in subset if m.frame_error_s is not None]) if subset else ""),
        })
    return rows


def _section_class_breakdown_rows(matches: list[EventMatch], gt_events: list[dict[str, str]]) -> list[dict]:
    tp = [m for m in matches if m.match_status == "TP" and m.gt_index is not None]
    grouped: dict[tuple[str, str], list[EventMatch]] = {}
    for m in tp:
        gt_class_name = str(gt_events[m.gt_index].get("class_name", ""))
        key = (m.section, gt_class_name)
        grouped.setdefault(key, []).append(m)

    rows: list[dict] = []
    for (section, gt_class), subset in sorted(grouped.items()):
        rows.append({
            "section": section,
            "gt_class_name": _class_display_name(gt_class),
            "gt_vehicle_group": _vehicle_group(gt_class),
            "tp_count": len(subset),
            "class_accuracy": _fmt(_subset_accuracy([m.class_correct for m in subset])),
            "lane_accuracy": _fmt(_subset_accuracy([m.lane_correct for m in subset])),
            "direction_accuracy": _fmt(_subset_accuracy([m.direction_correct for m in subset])),
            "crossing_time_mae_s": _fmt(mean([m.frame_error_s for m in subset if m.frame_error_s is not None]) if subset else ""),
        })
    return rows


def _section_lane_breakdown_rows(matches: list[EventMatch], gt_events: list[dict[str, str]]) -> list[dict]:
    tp = [m for m in matches if m.match_status == "TP" and m.gt_index is not None]
    grouped: dict[tuple[str, str], list[EventMatch]] = {}
    for m in tp:
        gt_lane_id = _normalize_lane_id(
            gt_events[m.gt_index].get("lane_id", ""),
            str(gt_events[m.gt_index].get("section", "")).strip(),
        )
        key = (m.section, gt_lane_id)
        grouped.setdefault(key, []).append(m)

    rows: list[dict] = []
    for (section, lane_id), subset in sorted(grouped.items(), key=lambda item: (item[0][0], item[0][1] or "ZZ")):
        rows.append({
            "section": section,
            "gt_lane_id": lane_id,
            "tp_count": len(subset),
            "class_accuracy": _fmt(_subset_accuracy([m.class_correct for m in subset])),
            "lane_accuracy": _fmt(_subset_accuracy([m.lane_correct for m in subset])),
            "direction_accuracy": _fmt(_subset_accuracy([m.direction_correct for m in subset])),
            "crossing_time_mae_s": _fmt(mean([m.frame_error_s for m in subset if m.frame_error_s is not None]) if subset else ""),
        })
    return rows


def validate_outputs(
    cross_section_csv: Path,
    manual_crossing_csv: Path,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    fps: float = 25.0,
    frame_tolerance: int = 10,
) -> dict:
    """验证断面事件与交通参数一致性，并保存报告数据。

    Returns:
        汇总指标字典，内容同时写入 validation_summary.csv。
    """
    raw_pred_events = _load_table(cross_section_csv)
    raw_gt_events = _load_table(manual_crossing_csv)
    output_dir.mkdir(parents=True, exist_ok=True)

    pred_events = _normalize_pred_events(raw_pred_events)
    pred_events = _dedupe_pred_events(pred_events)
    pred_sections = {str(r.get("section", "")).strip() for r in pred_events if str(r.get("section", "")).strip()}
    gt_events = _normalize_manual_events(raw_gt_events, allowed_sections=pred_sections or None)

    pred_max_frame = max((_to_int(r.get("frame_id")) for r in pred_events), default=None)
    gt_max_frame = max((_to_int(r.get("gt_frame_id")) for r in gt_events), default=None)
    overlap_max_frame = None
    if pred_max_frame is not None and gt_max_frame is not None:
        overlap_max_frame = min(pred_max_frame, gt_max_frame)
        gt_events = [r for r in gt_events if (_to_int(r.get("gt_frame_id")) or 0) <= overlap_max_frame]

    matches = _match_events(pred_events, gt_events, fps=fps, frame_tolerance=frame_tolerance)
    match_rows = _event_match_rows(matches)
    headway_rows = _headway_details(gt_events, pred_events, fps=fps, matches=matches)
    spacing_rows = _spacing_details(pred_events)
    anomalies = _anomaly_rows(pred_events, fps=fps)
    metric_breakdown_rows = _metric_breakdown_rows(matches, gt_events)
    section_class_rows = _section_class_breakdown_rows(matches, gt_events)
    section_lane_rows = _section_lane_breakdown_rows(matches, gt_events)
    summary = _summary(matches, headway_rows, spacing_rows, anomalies)
    summary.update(_breakdown_summary(matches, gt_events))
    summary.update({
        "pred_raw_count": len(raw_pred_events),
        "pred_dedup_count": len(pred_events),
        "manual_raw_count": len(raw_gt_events),
        "manual_in_scope_count": len(gt_events),
        "manual_excluded_count": max(len(raw_gt_events) - len(gt_events), 0),
        "pred_span_frames": pred_max_frame if pred_max_frame is not None else "",
        "manual_span_frames": gt_max_frame if gt_max_frame is not None else "",
        "overlap_max_frame": overlap_max_frame if overlap_max_frame is not None else "",
        "overlap_seconds": round((overlap_max_frame or 0) / fps, 2) if overlap_max_frame is not None else "",
    })

    generated_files: list[str] = []

    def write(name: str, rows: list[dict], fields: list[str]) -> None:
        _write_csv(output_dir / name, rows, fields)
        generated_files.append(name)

    summary_rows = [{"metric": key, "value": _fmt(value)} for key, value in summary.items()]
    write("validation_summary.csv", summary_rows, ["metric", "value"])
    write("event_matching_details.csv", match_rows, [
        "match_status", "pred_index", "gt_index", "section", "track_id",
        "lane_id", "class_name", "direction", "pred_frame_id", "gt_frame_id",
        "frame_error_s", "lane_correct", "direction_correct", "class_correct",
        "pred_lane_id", "gt_lane_id", "pred_class_name", "gt_class_name",
        "pred_direction", "gt_direction",
    ])
    write("headway_details.csv", headway_rows, [
        "section", "direction", "track_id", "gt_frame_id", "previous_gt_frame_id",
        "gt_headway_s", "pred_headway_s", "abs_error_s", "mape", "error_le_0_5s",
    ])
    write("spacing_consistency_details.csv", spacing_rows, [
        "section", "direction", "track_id", "frame_id", "previous_track_id",
        "previous_speed_kmh", "headway_s", "spacing_m", "expected_spacing_m",
        "spacing_error_m", "passed",
    ])
    write("anomaly_events.csv", anomalies, [
        "anomaly_type", "detail", "row_index", "frame_id", "timestamp_s",
        "section", "track_id", "class_name", "direction", "speed_kmh",
        "headway_s", "spacing_m",
    ])
    write("metric_breakdown.csv", metric_breakdown_rows, [
        "scope", "tp_count", "class_accuracy", "lane_accuracy", "direction_accuracy", "crossing_time_mae_s",
    ])
    write("section_class_breakdown.csv", section_class_rows, [
        "section", "gt_class_name", "gt_vehicle_group", "tp_count", "class_accuracy", "lane_accuracy", "direction_accuracy", "crossing_time_mae_s",
    ])
    write("section_lane_breakdown.csv", section_lane_rows, [
        "section", "gt_lane_id", "tp_count", "class_accuracy", "lane_accuracy", "direction_accuracy", "crossing_time_mae_s",
    ])

    log = {
        "cross_section_csv": str(cross_section_csv),
        "manual_crossing_csv": str(manual_crossing_csv),
        "output_dir": str(output_dir),
        "fps": fps,
        "frame_tolerance": frame_tolerance,
        "manual_event_count": len(gt_events),
        "pred_event_count": len(pred_events),
        "pred_raw_count": len(raw_pred_events),
        "manual_raw_count": len(raw_gt_events),
        "overlap_max_frame": overlap_max_frame,
        "generated_files": generated_files,
        "summary": summary,
        "notes": [
            "先对断面名称和车型名称做归一，再只保留算法实际覆盖的断面进行比较。",
            "车道号口径按人工标注约定归一：L1-L2 之间记为车道 1，以此类推。",
            "预测结果先按 track_id + section 折叠短间隔重复触发，再进入事件匹配。",
            "自动以预测与人工标注的共同帧段为重叠区间；超出重叠区间的人工记录不会计入 FN。",
            "方向字段按断面语义自动对齐：主断面对比 direction，右转断面对比 arrival_departure。",
            "车头间距按 previous_speed / 3.6 * headway_s 做衍生一致性校验。",
            "该验证报告用于人工复核集，不代表主办方隐藏测试评分。",
        ],
    }
    log_path = output_dir / "validation_log.json"
    log_path.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")
    generated_files.append(log_path.name)
    log["generated_files"] = generated_files
    log_path.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")

    return summary


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="基于人工复核集验证交通流算法输出。")
    parser.add_argument("--cross-section", type=Path, default=Path("outputs") / "cross_section.csv",
                        help="算法输出的 cross_section.csv")
    parser.add_argument("--manual-crossing", type=Path, required=True,
                        help="人工过线标注文件，支持 CSV 或 Excel，至少包含 gt_frame_id/section/class_name/direction")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR,
                        help="验证报告与明细输出目录")
    parser.add_argument("--export-normalized-manual", type=Path, default=None,
                        help="可选：把人工标注另存为规范 UTF-8 CSV 后再验证")
    parser.add_argument("--fps", type=float, default=25.0, help="视频 FPS")
    parser.add_argument("--frame-tolerance", type=int, default=10, help="事件匹配帧容差")
    return parser


def main() -> None:
    args = _build_parser().parse_args()
    manual_crossing = args.manual_crossing
    if args.export_normalized_manual is not None:
        events = export_normalized_manual_annotations(
            args.manual_crossing,
            args.export_normalized_manual,
        )
        manual_crossing = args.export_normalized_manual
        print(f"规范化人工标注已保存: {manual_crossing}（{len(events)} 条）")
    summary = validate_outputs(
        cross_section_csv=args.cross_section,
        manual_crossing_csv=manual_crossing,
        output_dir=args.output_dir,
        fps=args.fps,
        frame_tolerance=args.frame_tolerance,
    )
    print("=== 人工复核验证摘要 ===")
    for key, value in summary.items():
        print(f"{key}: {_fmt(value)}")
    print(f"验证明细已保存: {args.output_dir}")


if __name__ == "__main__":
    main()
