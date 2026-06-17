"""
轨迹相似度分组模块。

每15秒对已结束轨迹做一次批量分组：
  余弦相似度 + JS散度相似度 + 欧氏距离（三者同时超阈值才归组）

独立测试：
    python3 -m pytest tests/trajectory/test_traj_grouper.py -v
"""
from __future__ import annotations

import csv
import math
from collections import defaultdict
from pathlib import Path
from typing import Optional

import numpy as np

import sys as _sys
_sys.path.insert(0, str(Path(__file__).parents[2]))

from src.config.settings import (
    TRAJ_GROUP_INTERVAL_S,
    TRAJ_GROUP_COS_THRESH,
    TRAJ_GROUP_JSD_THRESH,
    TRAJ_GROUP_EUC_THRESH,
    TRAJ_GROUP_MIN_FRAMES,
    TRAJ_GROUP_CSV_PATH,
    ENTRANCE_TRAVEL_DIR,
    EXCEL_REPORT_PATH,
)

_CSV_FIELDS = [
    "window_start_s", "window_end_s", "entrance",
    "group_id", "turn_type", "class_type",
    "track_ids", "size",
]


class _UnionFind:
    def __init__(self, n: int) -> None:
        self._p = list(range(n))

    def find(self, x: int) -> int:
        while self._p[x] != x:
            self._p[x] = self._p[self._p[x]]
            x = self._p[x]
        return x

    def union(self, x: int, y: int) -> None:
        self._p[self.find(x)] = self.find(y)

    def groups(self) -> list[list[int]]:
        g: dict[int, list[int]] = defaultdict(list)
        for i in range(len(self._p)):
            g[self.find(i)].append(i)
        return list(g.values())


class TrajGrouper:
    """在线轨迹相似度分组器。

    用法：
        grouper = TrajGrouper()
        # 主循环内每帧：
        grouper.tick(timestamp_s)
        # 轨迹消失时：
        grouper.on_track_end(tid, points, cls, lane_id, lane_type, entrance)
        # 视频结束：
        grouper.finalize()
    """

    def __init__(
        self,
        cos_thresh: float = TRAJ_GROUP_COS_THRESH,
        jsd_thresh: float = TRAJ_GROUP_JSD_THRESH,
        euc_thresh: float = TRAJ_GROUP_EUC_THRESH,
        interval_s: float = TRAJ_GROUP_INTERVAL_S,
        min_frames: int   = TRAJ_GROUP_MIN_FRAMES,
        csv_path: Optional[Path | str] = TRAJ_GROUP_CSV_PATH,
        excel_path: Optional[Path | str] = EXCEL_REPORT_PATH,
        frame_width: int = 3840,
        frame_height: int = 2160,
    ) -> None:
        self.cos_thresh   = cos_thresh
        self.jsd_thresh   = jsd_thresh
        self.euc_thresh   = euc_thresh    # 比较时除以100
        self.interval_s   = interval_s
        self.min_frames   = min_frames
        self.frame_width  = frame_width
        self.frame_height = frame_height

        self._csv_path   = Path(csv_path)   if csv_path   else None
        self._excel_path = Path(excel_path) if excel_path else None

        # 待分组队列：entrance → list of (tid, points, cls, lane_id, lane_type)
        self._pending: dict[str, list[tuple]] = defaultdict(list)
        self._last_flush_s: float = 0.0
        self._group_counter: int  = 0
        self._all_records: list[dict] = []   # 内存缓存，供 Excel 导出

        if self._csv_path is not None:
            self._csv_path.parent.mkdir(parents=True, exist_ok=True)
            self._csv_fh = self._csv_path.open("w", newline="", encoding="utf-8")
            self._csv_writer = csv.DictWriter(self._csv_fh, fieldnames=_CSV_FIELDS)
            self._csv_writer.writeheader()
        else:
            self._csv_fh = None
            self._csv_writer = None

    def on_track_end(
        self,
        track_id: int,
        points: list[tuple[float, float]],
        cls: str,
        lane_id: Optional[int],
        lane_type: str,
        entrance: str,
    ) -> None:
        """轨迹结束时调用。帧数不足 min_frames 的片段丢弃。"""
        if len(points) < self.min_frames:
            return
        self._pending[entrance].append((track_id, points, cls, lane_id, lane_type))

    def tick(self, video_time_s: float) -> None:
        """每帧调用。满 interval_s 时触发批量分组。"""
        if video_time_s - self._last_flush_s >= self.interval_s:
            self._flush(self._last_flush_s, video_time_s)
            self._last_flush_s = video_time_s

    def finalize(self) -> None:
        """视频结束时强制 flush 剩余轨迹，写 Excel sheet。"""
        self._flush(self._last_flush_s, float("inf"))
        if self._csv_fh is not None:
            self._csv_fh.close()
        self._write_excel_sheet()

    def get_records(self) -> list[dict]:
        """返回所有分组记录（供外部写 Excel 时使用）。"""
        return list(self._all_records)

    def _compute_similarity(
        self,
        pts_a: list[tuple[float, float]],
        pts_b: list[tuple[float, float]],
        width: int,
        height: int,
    ) -> tuple[float, float, float]:
        """计算两条轨迹的三指标相似度。

        Returns:
            (cosine_sim, jsd_sim, euc_sim)
            - cosine_sim: [-1, 1]，越高越相似
            - jsd_sim:    [0, 1]，越高越相似
            - euc_sim:    [0, 1]，越高越相似（对角线归一化后的相似度）
        """
        # 对齐长度（零填充）
        n = max(len(pts_a), len(pts_b))
        arr_a = np.zeros((n, 2), dtype=np.float64)
        arr_b = np.zeros((n, 2), dtype=np.float64)
        arr_a[:len(pts_a)] = pts_a
        arr_b[:len(pts_b)] = pts_b

        # 余弦相似度（用首尾向量，对零填充鲁棒）
        vec_a = np.array(pts_a[-1]) - np.array(pts_a[0]) if len(pts_a) > 1 else np.zeros(2)
        vec_b = np.array(pts_b[-1]) - np.array(pts_b[0]) if len(pts_b) > 1 else np.zeros(2)
        norm_a = np.linalg.norm(vec_a)
        norm_b = np.linalg.norm(vec_b)
        if norm_a < 1e-9 or norm_b < 1e-9:
            cos_sim = 0.0
        else:
            cos_sim = float(np.dot(vec_a, vec_b) / (norm_a * norm_b))

        # JS散度相似度（归一化坐标→概率分布，使用直方图）
        norm_a_xy = arr_a / np.array([width, height], dtype=np.float64)
        norm_b_xy = arr_b / np.array([width, height], dtype=np.float64)
        hist_a = self._to_hist(norm_a_xy)
        hist_b = self._to_hist(norm_b_xy)
        try:
            from scipy.spatial.distance import jensenshannon
            jsd = float(jensenshannon(hist_a, hist_b))
        except Exception:
            jsd = 0.5
        jsd_sim = max(0.0, 1.0 - jsd)

        # 欧氏距离相似度（对角线归一化，转为[0,1]相似度）
        diag = math.sqrt(width ** 2 + height ** 2)
        diff = arr_a - arr_b
        euc_norm = float(np.sqrt((diff ** 2).sum(axis=1)).mean()) / diag
        euc_sim = 1.0 / (1.0 + euc_norm)

        return cos_sim, jsd_sim, euc_sim

    @staticmethod
    def _to_hist(norm_xy: np.ndarray, bins: int = 20) -> np.ndarray:
        """将归一化坐标序列转为 bins×bins 的展平概率分布。"""
        x = np.clip(norm_xy[:, 0], 0.0, 1.0 - 1e-9)
        y = np.clip(norm_xy[:, 1], 0.0, 1.0 - 1e-9)
        idx = (x * bins).astype(int) * bins + (y * bins).astype(int)
        hist = np.bincount(idx, minlength=bins * bins).astype(np.float64)
        total = hist.sum()
        if total > 0:
            hist /= total
        else:
            hist[:] = 1.0 / (bins * bins)
        return hist

    def _infer_turn_type(
        self,
        points: list[tuple[float, float]],
        lane_type: str,
        entrance: str,
    ) -> str:
        """推断转向类型。lane_type 含 motor/non-motor，转向由几何角度决定。"""
        if lane_type and "non" in lane_type.lower():
            return "non_motor"

        if len(points) < 2:
            return "unknown"

        start = points[0]
        end   = points[-1]
        dx = end[0] - start[0]
        dy = end[1] - start[1]
        if abs(dx) < 1e-6 and abs(dy) < 1e-6:
            return "unknown"

        ref = ENTRANCE_TRAVEL_DIR.get(entrance, (0.0, 1.0))
        dot   =  dx * ref[0] + dy * ref[1]
        cross =  dx * ref[1] - dy * ref[0]
        angle_deg = math.degrees(math.atan2(cross, dot))

        if abs(angle_deg) < 30:
            return "straight"
        elif -90 < angle_deg < -30:
            return "left_turn"
        elif 30 < angle_deg < 90:
            return "right_turn"
        else:
            return "wrong_way"

    def _flush(self, win_start: float, win_end: float) -> None:
        """对每个进口的 pending 队列做一次分组并写出结果。"""
        for entrance, items in self._pending.items():
            if not items:
                continue
            n = len(items)
            uf = _UnionFind(n)

            pairs: list[tuple[float, int, int]] = []
            for i in range(n):
                for j in range(i + 1, n):
                    pts_i = items[i][1]
                    pts_j = items[j][1]
                    cos_s, jsd_s, euc_s = self._compute_similarity(
                        pts_i, pts_j, self.frame_width, self.frame_height
                    )
                    if (cos_s >= self.cos_thresh
                            and jsd_s >= self.jsd_thresh
                            and euc_s >= self.euc_thresh):
                        pairs.append((cos_s, i, j))

            for _, i, j in sorted(pairs, reverse=True):
                uf.union(i, j)

            win_end_str = win_end if win_end != float("inf") else win_start + self.interval_s
            for group_indices in uf.groups():
                self._group_counter += 1
                gid = f"G{self._group_counter:04d}"

                members = [items[k] for k in group_indices]
                cls_counts: dict[str, int] = defaultdict(int)
                for _, _, cls, _, _ in members:
                    cls_counts[cls] += 1
                dominant_cls = max(cls_counts, key=lambda k: cls_counts[k])

                _, rep_pts, _, _, rep_lane_type = members[0]
                turn = self._infer_turn_type(rep_pts, rep_lane_type, entrance)

                track_ids_str = ",".join(str(items[k][0]) for k in group_indices)
                row = {
                    "window_start_s": round(win_start, 1),
                    "window_end_s":   round(float(win_end_str), 1),
                    "entrance":       entrance,
                    "group_id":       gid,
                    "turn_type":      turn,
                    "class_type":     dominant_cls,
                    "track_ids":      track_ids_str,
                    "size":           len(group_indices),
                }
                if self._csv_writer is not None:
                    self._csv_writer.writerow(row)
                self._all_records.append(row)

        self._pending.clear()

    def _write_excel_sheet(self) -> None:
        """向 traffic_report.xlsx 追加"轨迹分组" sheet。"""
        if not self._all_records or self._excel_path is None:
            return
        try:
            import openpyxl
        except ImportError:
            print("[TrajGrouper] openpyxl 未安装，跳过 Excel 导出")
            return

        if self._excel_path.exists():
            wb = openpyxl.load_workbook(self._excel_path)
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        sheet_name = "轨迹分组"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        headers_cn = {
            "window_start_s": "窗口开始(秒)",
            "window_end_s":   "窗口结束(秒)",
            "entrance":       "进口",
            "group_id":       "分组ID",
            "turn_type":      "转向类型",
            "class_type":     "车辆类型",
            "track_ids":      "track_id列表",
            "size":           "车辆数",
        }
        ws.append([headers_cn[f] for f in _CSV_FIELDS])
        for rec in self._all_records:
            ws.append([rec[f] for f in _CSV_FIELDS])

        wb.save(self._excel_path)
        print(f"[TrajGrouper] Excel 写入完成：{self._excel_path}（sheet=轨迹分组，{len(self._all_records)} 行）")
