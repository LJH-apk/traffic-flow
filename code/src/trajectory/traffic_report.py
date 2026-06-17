"""轨迹、断面事件和 Excel 报表输出。"""
from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

import numpy as np

from src.config.settings import (
    EXCEL_REPORT_PATH,
    QUEUE_GAP_M,
    QUEUE_SPEED_THRESH_KMH,
    SECTION_ROAD_LENGTH_M,
    VEHICLE_LENGTHS_M,
)
from src.trajectory.plate_recognizer import PlateBox


def rewrite_dict_csv(path: Path, fields: list[str], rows: list[dict]) -> None:
    with Path(path).open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def backfill_plates(
    traj_records: list[dict],
    cross_events: list[dict],
    plate_cache: dict[int, tuple[str, PlateBox | None]],
) -> int:
    plate_by_track_id = {
        int(track_id): plate
        for track_id, (plate, _) in plate_cache.items()
        if plate
    }
    filled = 0
    for rows in (traj_records, cross_events):
        for row in rows:
            try:
                track_id = int(row.get("track_id"))
            except (TypeError, ValueError):
                continue
            plate = plate_by_track_id.get(track_id)
            if plate and not row.get("plate"):
                row["plate"] = plate
                filled += 1
    return filled


def write_vehicle_stats_row(
    writer: csv.writer,
    track_id: int,
    stats: dict | None,
    lane_id: int | str | None,
    min_samples: int,
) -> None:
    if not stats or stats["n_samples"] < min_samples:
        return
    writer.writerow([
        track_id,
        stats["first_frame"],
        stats["last_frame"],
        lane_id,
        round(stats["avg_kmh"], 1),
        round(stats["max_kmh"], 1),
        round(stats["min_kmh"], 1),
        stats["n_samples"],
    ])


def export_excel(
    traj_records: list[dict],
    cross_events: list[dict],
    traj_fields: list[str],
    cross_fields: list[str],
    output_path: Path = EXCEL_REPORT_PATH,
) -> None:
    try:
        import openpyxl
    except ImportError:
        print("[Excel] 未找到 openpyxl，跳过 Excel 导出（pip install openpyxl）")
        return

    field_names = {
        "frame_id": "帧编号",
        "timestamp_s": "时间戳(秒)",
        "section": "断面名称",
        "arrival_departure": "到达/离去",
        "track_id": "车辆ID",
        "class_name": "车辆类型",
        "vehicle_category": "车辆性质",
        "lane_id": "车道编号",
        "color": "车身颜色",
        "direction": "行驶方向",
        "speed_kmh": "速度(km/h)",
        "headway_s": "车头时距(秒)",
        "spacing_m": "车头间距(米)",
        "cx": "中心X",
        "cy": "中心Y",
        "x1": "左上X",
        "y1": "左上Y",
        "x2": "右下X",
        "y2": "右下Y",
        "plate": "车牌号",
    }

    def headers(fields: list[str]) -> list[str]:
        return [field_names.get(field, field) for field in fields]

    wb = openpyxl.Workbook()

    ws_cross = wb.active
    ws_cross.title = "断面过车"
    ws_cross.append(headers(cross_fields))
    for event in cross_events:
        ws_cross.append([event.get(field, "") for field in cross_fields])

    ws_traj = wb.create_sheet("车辆轨迹")
    ws_traj.append(headers(traj_fields))
    for record in traj_records:
        ws_traj.append([record.get(field, "") for field in traj_fields])

    ws_flow = wb.create_sheet("流量统计")
    ws_flow.append(["统计起始(秒)", "统计结束(秒)", "断面名称", "车辆性质", "过车数", "流量(辆/分钟)"])
    flow_count: dict[tuple, int] = defaultdict(int)
    for event in cross_events:
        window = int(float(event.get("timestamp_s", 0)) // 60)
        flow_count[(window, event.get("section", ""), event.get("vehicle_category", ""))] += 1
    for (window, section, category), count in sorted(flow_count.items()):
        ws_flow.append([window * 60, (window + 1) * 60, section, category, count, round(count / 1.0, 2)])

    ws_occ = wb.create_sheet("空间占有率")
    ws_occ.append([
        "统计起始(秒)", "统计结束(秒)", "车道编号",
        "平均在场车辆数", "平均占用总长度(米)", "平均空间占有率(%)",
    ])
    occ_snapshot: dict[tuple, list] = defaultdict(list)
    for record in traj_records:
        lane_id = record.get("lane_id", "")
        if not _valid_lane_id(lane_id):
            continue
        timestamp = int(float(record.get("timestamp_s", 0)))
        occ_snapshot[(timestamp, lane_id)].append(record)

    occ_window: dict[tuple, list] = defaultdict(list)
    for (timestamp, lane_id), records in occ_snapshot.items():
        deduped = _dedupe_snapshot(records)
        total_len = sum(VEHICLE_LENGTHS_M.get(record.get("class_name", "car"), 4.5) for record in deduped)
        occ_pct = round(min(100.0, total_len / SECTION_ROAD_LENGTH_M * 100), 1)
        occ_window[(timestamp // 5, lane_id)].append((len(deduped), total_len, occ_pct))

    for (window, lane_id) in sorted(occ_window.keys(), key=lambda item: (item[0], str(item[1]))):
        samples = occ_window[(window, lane_id)]
        n_samples = len(samples)
        avg_count = round(sum(value[0] for value in samples) / n_samples, 1)
        avg_len = round(sum(value[1] for value in samples) / n_samples, 1)
        avg_occ = round(sum(value[2] for value in samples) / n_samples, 1)
        ws_occ.append([window * 5, (window + 1) * 5, lane_id, avg_count, avg_len, avg_occ])

    ws_queue = wb.create_sheet("排队长度")
    ws_queue.append([
        "统计起始(秒)", "统计结束(秒)", "车道编号",
        "排队车辆数", "排队长度(米)", "平均排队车速(km/h)",
    ])
    motor_classes = {"car", "bus", "truck"}
    queue_bucket: dict[tuple, list] = defaultdict(list)
    for record in traj_records:
        if record.get("class_name") not in motor_classes:
            continue
        if record.get("lane_type") != "motor":
            continue
        speed_raw = record.get("speed_kmh")
        if speed_raw is None or speed_raw == "":
            continue
        speed = float(speed_raw)
        if 0.0 <= speed <= QUEUE_SPEED_THRESH_KMH:
            lane_id = record.get("lane_id", "")
            if not _valid_lane_id(lane_id):
                continue
            timestamp = int(float(record.get("timestamp_s", 0)))
            queue_bucket[(timestamp, lane_id)].append(record)

    for (timestamp, lane_id) in sorted(queue_bucket.keys(), key=lambda item: (item[0], str(item[1]))):
        records = _dedupe_snapshot(queue_bucket[(timestamp, lane_id)])
        count = len(records)
        avg_speed = round(sum(float(record.get("speed_kmh", 0) or 0) for record in records) / count, 1)
        total_body = sum(VEHICLE_LENGTHS_M.get(record.get("class_name", "car"), 4.5) for record in records)
        queue_len = round(total_body + max(0, count - 1) * QUEUE_GAP_M, 1)
        ws_queue.append([timestamp, timestamp + 1, lane_id, count, queue_len, avg_speed])

    wb.save(output_path)
    print(f"[Excel] 报表已导出: {output_path}")


def _valid_lane_id(value) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, str):
        raw = value.strip().upper()
        if raw == "OPPOSITE":
            return False
        if not raw.isdigit():
            return False
    try:
        return not np.isnan(value)
    except TypeError:
        return True


def _bbox_iou(a: dict, b: dict) -> float:
    ax1, ay1, ax2, ay2 = (float(a.get(key, 0) or 0) for key in ("x1", "y1", "x2", "y2"))
    bx1, by1, bx2, by2 = (float(b.get(key, 0) or 0) for key in ("x1", "y1", "x2", "y2"))
    ix1, iy1 = max(ax1, bx1), max(ay1, by1)
    ix2, iy2 = min(ax2, bx2), min(ay2, by2)
    iw, ih = max(0.0, ix2 - ix1), max(0.0, iy2 - iy1)
    inter = iw * ih
    area_a = max(0.0, ax2 - ax1) * max(0.0, ay2 - ay1)
    area_b = max(0.0, bx2 - bx1) * max(0.0, by2 - by1)
    denom = area_a + area_b - inter
    return inter / denom if denom > 0 else 0.0


def _center_dist(a: dict, b: dict) -> float:
    return float(np.hypot(
        float(a.get("cx", 0) or 0) - float(b.get("cx", 0) or 0),
        float(a.get("cy", 0) or 0) - float(b.get("cy", 0) or 0),
    ))


def _dedupe_snapshot(records: list[dict]) -> list[dict]:
    kept: list[dict] = []
    for record in sorted(records, key=lambda item: VEHICLE_LENGTHS_M.get(item.get("class_name", "car"), 4.5)):
        if any(_bbox_iou(record, old) >= 0.85 or _center_dist(record, old) <= 5.0 for old in kept):
            continue
        kept.append(record)
    return kept
