#!/usr/bin/env python3
"""将三进口全量 CSV 整理为一个 xlsx，中文正常显示。"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parents[1]
OUTPUTS = ROOT / "outputs"
OUT_PATH = OUTPUTS / "三进口全量数据.xlsx"

ENTRANCES = [
    ("北进口", "north"),
    ("东进口", "east"),
    ("南进口", "south"),
]

# ── 样式 ──────────────────────────────────────────────────────────
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
CELL_ALIGN = Alignment(vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
ENTRANCE_FILLS = {
    "北进口": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "东进口": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
    "南进口": PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid"),
}
# 交替行底色（无进口色时使用）
ROW_FILL_ODD = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")


def _read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _write_sheet(ws, headers, rows, entrance_col=None):
    """将 headers + rows 写入 worksheet，自动列宽和样式。"""
    # 写表头
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # 写数据
    for ri, row in enumerate(rows, start=2):
        entrance_name = row.get(entrance_col, "") if entrance_col else ""
        fill = ENTRANCE_FILLS.get(entrance_name) or (ROW_FILL_ODD if ri % 2 == 0 else None)
        for ci, h in enumerate(headers, start=1):
            val = row.get(h, "")
            # 尝试转数字
            if isinstance(val, str):
                try:
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (ValueError, OverflowError):
                    pass
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

    # 自动列宽
    for ci, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for ri in range(2, min(len(rows) + 2, 100)):  # 取样前100行
            cell_val = ws.cell(row=ri, column=ci).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 36)

    # 冻结首行
    ws.freeze_panes = "A2"


def _load_all(fname_prefix):
    """加载三进口同名文件，合并为一个列表，每条记录附加入口列。"""
    all_rows = []
    for entrance, suffix in ENTRANCES:
        path = OUTPUTS / f"{fname_prefix}_{suffix}_full.csv"
        if not path.exists():
            continue
        rows = _read_csv(path)
        for row in rows:
            row["进口"] = entrance
        all_rows.extend(rows)
    return all_rows


def main():
    print("导出三进口数据 → xlsx …")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认 sheet

    # ── Sheet 1: 断面过车事件 ──
    print("[1/5] 断面过车事件 …")
    cross = _load_all("cross_section")
    headers_cross = ["进口", "frame_id", "timestamp_s", "section", "arrival_departure",
                     "track_id", "plate", "class_name", "lane_id", "vehicle_category",
                     "color", "direction", "speed_kmh", "headway_s", "spacing_m"]
    # 只保留实际存在的列
    if cross:
        existing = [h for h in headers_cross if h in cross[0]]
        ws1 = wb.create_sheet("断面过车事件")
        _write_sheet(ws1, existing, cross, entrance_col="进口")
    else:
        ws1 = wb.create_sheet("断面过车事件（空）")

    # ── Sheet 2: 轨迹数据 ──
    print("[2/5] 轨迹数据 …")
    traj = _load_all("trajectory")
    headers_traj = ["进口", "frame_id", "timestamp_s", "track_id", "class_name",
                    "lane_id", "lane_type", "cx", "cy", "x1", "y1", "x2", "y2",
                    "speed_kmh", "plate"]
    if traj:
        existing = [h for h in headers_traj if h in traj[0]]
        ws2 = wb.create_sheet("轨迹数据")
        _write_sheet(ws2, existing, traj, entrance_col="进口")
    else:
        wb.create_sheet("轨迹数据（空）")

    # ── Sheet 3: 车辆统计 ──
    print("[3/5] 车辆统计 …")
    stats = _load_all("vehicle_stats")
    headers_stats = ["进口", "track_id", "first_frame", "last_frame", "lane_id",
                     "avg_speed_kmh", "max_speed_kmh", "min_speed_kmh", "n_samples"]
    if stats:
        existing = [h for h in headers_stats if h in stats[0]]
        ws3 = wb.create_sheet("车辆统计")
        _write_sheet(ws3, existing, stats, entrance_col="进口")
    else:
        wb.create_sheet("车辆统计（空）")

    # ── Sheet 4: 轨迹分组 ──
    print("[4/5] 轨迹分组 …")
    groups = _load_all("trajectory_groups")
    headers_grp = ["进口", "window_start_s", "window_end_s", "entrance", "group_id",
                   "turn_type", "class_type", "track_ids", "size"]
    if groups:
        existing = [h for h in headers_grp if h in groups[0]]
        ws4 = wb.create_sheet("轨迹分组")
        _write_sheet(ws4, existing, groups, entrance_col="进口")
    else:
        wb.create_sheet("轨迹分组（空）")

    # ── Sheet 5: 汇总统计 ──
    print("[5/5] 汇总统计 …")
    ws5 = wb.create_sheet("汇总统计")

    summary_rows = []
    for entrance, suffix in ENTRANCES:
        cross_path = OUTPUTS / f"cross_section_{suffix}_full.csv"
        traj_path = OUTPUTS / f"trajectory_{suffix}_full.csv"
        stats_path = OUTPUTS / f"vehicle_stats_{suffix}_full.csv"

        n_events = n_traj = n_vehicles = 0
        avg_speed = 0.0
        class_counts = {}

        if cross_path.exists():
            cross_rows = _read_csv(cross_path)
            n_events = len(cross_rows)
            speeds = []
            for r in cross_rows:
                cls = r.get("class_name", "未知")
                class_counts[cls] = class_counts.get(cls, 0) + 1
                try:
                    s = float(r.get("speed_kmh", 0))
                    if s > 0:
                        speeds.append(s)
                except (ValueError, TypeError):
                    pass
            avg_speed = sum(speeds) / len(speeds) if speeds else 0.0

        if traj_path.exists():
            n_traj = len(_read_csv(traj_path))

        if stats_path.exists():
            track_ids = set()
            for r in _read_csv(stats_path):
                tid = r.get("track_id", "")
                if tid:
                    track_ids.add(tid)
            n_vehicles = len(track_ids)

        summary_rows.append({
            "进口": entrance,
            "过车事件数": n_events,
            "轨迹记录行": n_traj,
            "唯一车辆数": n_vehicles,
            "平均速度(km/h)": round(avg_speed, 1),
            "车型分布": "  ".join(f"{k}:{v}" for k, v in sorted(class_counts.items())),
        })

    hdr5 = list(summary_rows[0].keys()) if summary_rows else ["进口", "过车事件数"]
    _write_sheet(ws5, hdr5, summary_rows)

    # ── 保存 ──
    wb.save(OUT_PATH)
    print(f"\n✓ 导出完成: {OUT_PATH}")
    print(f"  共 {len(wb.sheetnames)} 个 Sheet: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
