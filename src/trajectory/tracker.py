"""
轨迹数据提取模块：ByteTrack 跟踪 + 车牌识别 + CSV 输出。

输出字段（trajectory.csv）::

    frame_id, timestamp_s, track_id, class_name, cx, cy, x1, y1, x2, y2, speed_kmh, plate

运行示例::

    python3 -u src/trajectory/tracker.py
"""
# 确保从任意目录运行时均可找到项目根（src 包）
import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).parents[2]))

import csv
import re
import time
from pathlib import Path
from typing import Optional

import cv2
import numpy as np
from ultralytics import YOLO

from src.config.settings import (
    VEHICLE_CLASSES,
    DEVICE,
    CONF_THRESH,
    VIDEO_PATH,
    OUTPUT_DIR,
    MODEL_DIR,
    MODEL_NAME,
    TRAJ_SAMPLE_FPS,
    TRAJ_CSV_PATH,
    SECTION_LINES,
    SECTION_LINES_MAP,
    ENTRANCE_ALIASES,
    HOMOGRAPHY_MATRIX,
    PIXELS_PER_METER,
    CROSS_SECTION_CSV_PATH,
    SPEED_WINDOW_FRAMES,
    SPEED_MIN_DIST_M,
    SPEED_MIN_SAMPLES,
    VEHICLE_STATS_CSV_PATH,
    VEHICLE_LENGTHS_M,
    SECTION_ROAD_LENGTH_M,
    QUEUE_SPEED_THRESH_KMH,
    QUEUE_GAP_M,
    EXCEL_REPORT_PATH,
    TRAJ_GROUP_INTERVAL_S,
    TRAJ_GROUP_CSV_PATH,
)
from src.cross_section.zebra_detector import ZebraDetector
from src.cross_section.counter import CrossSectionDetector
from src.cross_section.lane_detector import LaneDetector
from src.cross_section.speed_estimator import SpeedEstimator
from src.cross_section.lane_calibration import get_calibration
from src.cross_section.section_calibration import load_all_section_lines, load_section_lines
from src.trajectory.traj_grouper import TrajGrouper
from src.utils.video_io import open_video, video_meta, make_writer, iter_frames, AsyncWriter
from src.utils.visualization import draw_boxes, put_fps_text, put_text

# ── 运行时配置 ────────────────────────────────────────────────────────────────
_MODEL       = MODEL_DIR / MODEL_NAME          # 推理模型路径
_START_FRAME = 0                               # 起始帧号（含），0 = 视频开头
_END_FRAME   = 9000                            # 终止帧号（不含），None = 视频结尾
_GRACE_FRAMES = 10                              # track消失后保留帧数，防止碎片化
_OUTPUT_VID  = OUTPUT_DIR / "trajectory.mp4"  # 带轨迹标注的输出视频
_LIVE_PREVIEW_FPS = 18.0                       # Web 实时预览发布帧率


class _TrackerOutputLock:
    """跨进程输出锁，避免 dashboard 和命令行同时写 outputs。"""

    def __init__(self, path: Path) -> None:
        self.path = Path(path)
        self._fh = None

    def __enter__(self):
        import fcntl

        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._fh = self.path.open("w", encoding="utf-8")
        try:
            fcntl.flock(self._fh, fcntl.LOCK_EX | fcntl.LOCK_NB)
        except BlockingIOError as exc:
            self._fh.close()
            self._fh = None
            raise RuntimeError("已有检测任务正在写 outputs，请先停止 dashboard 检测或等待完成") from exc
        self._fh.write(f"{time.time():.3f}\n")
        self._fh.flush()
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        import fcntl

        if self._fh is None:
            return
        fcntl.flock(self._fh, fcntl.LOCK_UN)
        self._fh.close()
        self._fh = None

    def __del__(self) -> None:
        self.__exit__(None, None, None)


def _resolve_section_lines(video_path) -> list:
    """从视频文件名推断进口名，优先返回手动标注 sections.json。"""
    stem = Path(video_path).stem
    for alias, canonical in ENTRANCE_ALIASES.items():
        if alias in stem:
            lines = load_section_lines(canonical)
            if lines:
                print(f"[断面] 识别到进口：{canonical}，加载 {len(lines)} 条断面线")
                return lines
    lines = load_all_section_lines() or SECTION_LINES
    print(f"[断面] 未识别进口名（文件：{Path(video_path).name}），加载全部 {len(lines)} 条断面线")
    return lines
_TEST_VIDEO  = "北进口_20260420075959至20260420081500.mp4"

# 车牌正则表达式匹配：省份简称 + 字母 + 5位字母/数字
# 第1位：中文省份简称
# 第2位：A-Z（发牌城市代码）
# 第3-7位：A-Z 或 0-9（流水号，新能源末位可为字母，共5位）
_PROVINCE = "京津沪渝冀豫云辽黑湘皖鲁新苏浙赣鄂桂甘晋蒙陕吉闽贵粤川青藏琼宁"
# 完整车牌：省份简称 + 城市码(A-Z) + 5位流水号
_PLATE_FULL = re.compile(rf"[{_PROVINCE}][A-Z][A-Z0-9]{{5}}")
# 降级正则：省份字符被 OCR 误读时，从末尾取城市码 + 5位流水号（共6字符）
# 不加边界限制，直接取字符串最后6位，避免 'MA8R5Z9' 中 'A' 被前置 'M' 阻断
_PLATE_BODY = re.compile(r"[A-Z][A-Z0-9]{5}$")
_PLATE_CONF_THRESH = 0.85  # HyperLPR3 置信度阈值，低于此值的结果丢弃
_VOTE_MIN = 3              # 至少累积多少次高置信结果才锁定
_VOTE_MAX = 7              # 每辆车最多保留多少票（防止旧错误票污染）
_PLATE_BODY_MIN_VOTES = 3
_PLATE_BODY_MIN_RATIO = 0.50
_PLATE_PROVINCE_MIN_VOTES = 2
_PLATE_PROVINCE_MIN_RATIO = 0.60
_PLATE_MIN_BOTTOM_RATIO = 0.55  # 车框底部进入画面下半部后，脱敏马赛克才更可能解除
_PLATE_MIN_BOX_W = 160          # 过小车辆通常处于远处/低清晰度区域
_PLATE_MIN_BOX_H = 110
_PLATE_ROI_TOP_RATIO = 0.45     # 只在车辆下半部分找车牌，减少车身文字误识别
_PLATE_VEHICLE_CLASSES = {"car", "bus", "truck"}


def _rel_to_abs(
    rel_box: tuple[float, float, float, float],
    x1: int, y1: int, x2: int, y2: int,
) -> tuple[int, int, int, int]:
    """将相对坐标，转换为当前帧的绝对像素坐标。"""
    rx1, ry1, rx2, ry2 = rel_box
    vw, vh = x2 - x1, y2 - y1
    return (x1 + int(rx1 * vw), y1 + int(ry1 * vh),
            x1 + int(rx2 * vw), y1 + int(ry2 * vh))


# ─────────────────────────────────────────────────────────────────────────────
class PlateRecognizer:
    """车牌识别器：使用 HyperLPR3 检测并识别车牌，多帧投票后锁定结果。

    缓存值为 (plate_str, rel_box)，其中 rel_box 是车牌在车辆 bbox 内的相对坐标，
    供上层绘制真实车牌框使用。

    Attributes:
        _cache:   track_id -> (plate_str, rel_box)，已锁定的识别结果。
        _pending: track_id -> list of (plate_str, conf, rel_box)，待投票的候选列表。
        _enabled: False 表示 hyperlpr3 不可用，recognize() 直接返回空结果。
    """

    # 投票条目类型：(车牌字符串, 置信度, 相对坐标)
    _VoteEntry = tuple[str, float, tuple[float, float, float, float] | None]

    def __init__(self) -> None:
        self._cache:   dict[int, tuple[str, tuple[float, float, float, float] | None]] = {}
        self._pending: dict[int, list[PlateRecognizer._VoteEntry]] = {}
        self._catcher = None
        self._enabled = False

        try:
            import hyperlpr3 as lpr3  # type: ignore
            self._catcher = lpr3.LicensePlateCatcher()
            self._enabled = True
            print("[PlateRecognizer] HyperLPR3 加载成功")
        except ImportError:
            print("[PlateRecognizer] 未找到 hyperlpr3，车牌识别已禁用")

    def recognize(
        self,
        frame: np.ndarray,
        track_id: int,
        box_xyxy: tuple[int, int, int, int],
    ) -> tuple[str, tuple[float, float, float, float] | None]:
        """识别车牌，累积多帧投票后锁定最优结果。

        Args:
            frame:    完整 BGR 帧。
            track_id: 跟踪 ID。
            box_xyxy: 车辆检测框 (x1, y1, x2, y2)。

        Returns:
            (plate_str, rel_box)；rel_box 为车牌在车辆 bbox 内的相对坐标
            (rx1, ry1, rx2, ry2)，未锁定时返回 ("", None)。
        """
        if track_id in self._cache:
            return self._cache[track_id]
        if not self._enabled:
            return "", None

        fh, fw = frame.shape[:2]
        x1, y1, x2, y2 = box_xyxy
        cx1, cy1 = max(0, x1), max(0, y1)
        cx2, cy2 = min(fw, x2), min(fh, y2)
        crop_w, crop_h = cx2 - cx1, cy2 - cy1
        if crop_w <= 0 or crop_h <= 0:
            return "", None

        if cy2 / fh < _PLATE_MIN_BOTTOM_RATIO:
            return "", None
        if crop_w < _PLATE_MIN_BOX_W or crop_h < _PLATE_MIN_BOX_H:
            return "", None

        roi_y1 = cy1 + int(crop_h * _PLATE_ROI_TOP_RATIO)
        crop = frame[roi_y1:cy2, cx1:cx2]
        roi_h = cy2 - roi_y1
        if roi_h <= 0:
            return "", None

        try:
            results = self._catcher(crop)
        except Exception:  # noqa: BLE001
            return "", None

        for item in (results or []):
            text, conf = item[0], float(item[1])
            if conf < _PLATE_CONF_THRESH:
                continue
            plate = self._match_plate(text)
            if not plate:
                continue
            # item[3] 是车牌在 crop 内的绝对像素坐标，转为相对比例存储，
            # 使得后续每帧可根据当前 bbox 还原绝对坐标，车牌框跟随车辆移动。
            px1, py1, px2, py2 = item[3]
            rel_box = (
                px1 / crop_w,
                (roi_y1 - cy1 + py1) / crop_h,
                px2 / crop_w,
                (roi_y1 - cy1 + py2) / crop_h,
            )

            votes = self._pending.setdefault(track_id, [])
            votes.append((plate, conf, rel_box))
            if len(votes) > _VOTE_MAX:
                votes.pop(0)

        votes = self._pending.get(track_id, [])
        if len(votes) >= _VOTE_MIN:
            result, stable = self._tally(votes)
            if stable or len(votes) >= _VOTE_MAX:
                self._cache[track_id] = result
                del self._pending[track_id]
                return result

        return "", None

    @staticmethod
    def _tally(
        votes: list[tuple[str, float, tuple[float, float, float, float] | None]],
    ) -> tuple[tuple[str, tuple[float, float, float, float] | None], bool]:
        """对多帧投票结果取多数，省份字符与车牌主体分别投票。

        主体稳定即可输出；省份字更容易误识别，必须满足更严格的一致性。
        返回值第二项表示省份和主体是否都已稳定，可安全锁定完整车牌。
        """
        from collections import Counter

        provinces: list[str] = []
        bodies: list[str] = []
        for plate, _, _ in votes:
            province, body = PlateRecognizer._split_plate(plate)
            if province:
                provinces.append(province)
            if body:
                bodies.append(body)

        body = ""
        body_stable = False
        if bodies:
            body, body_votes = Counter(bodies).most_common(1)[0]
            body_stable = (
                body_votes >= _PLATE_BODY_MIN_VOTES
                and body_votes / len(bodies) >= _PLATE_BODY_MIN_RATIO
            )

        province = ""
        province_stable = False
        if provinces:
            province, province_votes = Counter(provinces).most_common(1)[0]
            province_stable = (
                province_votes >= _PLATE_PROVINCE_MIN_VOTES
                and province_votes / len(provinces) >= _PLATE_PROVINCE_MIN_RATIO
            )

        plate = (province + body) if (province_stable and body) else body
        rel_box = max(votes, key=lambda v: v[1])[2]  # 取置信度最高的 box
        return (plate, rel_box), (body_stable and province_stable)

    @staticmethod
    def _split_plate(plate: str) -> tuple[str, str]:
        """拆分为省份字符和 6 位主体，兼容缺省份的降级结果。"""
        if len(plate) >= 7 and plate[0] in _PROVINCE:
            return plate[0], plate[1:7]
        if len(plate) >= 6:
            return "", plate[-6:]
        return "", ""

    @staticmethod
    def _match_plate(text: str) -> str:
        """从 OCR 文字中提取车牌号。

        优先匹配完整格式（省份+城市码+5位），
        降级匹配城市码+5位（省份被误读时）。
        """
        clean = text.replace(" ", "").upper()
        m = _PLATE_FULL.search(clean)
        if m:
            return m.group()
        m = _PLATE_BODY.search(clean)
        if m:
            return m.group()
        return ""


# ─────────────────────────────────────────────────────────────────────────────
class TrajectoryTracker:
    """ByteTrack 车辆跟踪器，按固定频率采样轨迹并写出 CSV。

    每隔 ``sample_interval`` 帧记录一次各车辆的中心坐标、检测框和车牌，
    最终输出结构化 CSV 文件，供后续统计分析使用。

    Attributes:
        model_path:      YOLO 权重路径。
        device:          推理设备。
        conf:            置信度阈值。
        sample_interval: 每隔多少帧采样一次（根据视频 FPS 和 TRAJ_SAMPLE_FPS 计算）。
    """

    _CSV_FIELDS = [
        "frame_id", "timestamp_s", "track_id", "class_name", "lane_id", "lane_type",
        "cx", "cy", "x1", "y1", "x2", "y2", "speed_kmh", "plate",
    ]
    _CROSS_CSV_FIELDS = CrossSectionDetector.CSV_FIELDS

    def __init__(
        self,
        model_path: str | Path = _MODEL,
        device: str = DEVICE,
        conf: float = CONF_THRESH,
    ) -> None:
        """初始化跟踪器，加载 YOLO 模型和车牌识别器。

        Args:
            model_path: YOLO 权重文件路径。
            device:     推理设备（mps / cpu / cuda）。
            conf:       置信度阈值。
        """
        self.model_path = Path(model_path)
        self.device     = device
        self.conf       = conf
        self._model     = YOLO(str(self.model_path))
        self._model.to(device)
        self._plate_rec = PlateRecognizer()
        print(f"[Tracker] 模型: {self.model_path.name}  设备: {device}")

    @staticmethod
    def _build_lane_overlay(
        height: int, width: int,
        lanes: dict[int, list[tuple[int, int]]],
    ) -> np.ndarray:
        """根据标注车道线点集生成半透明 overlay（启动时一次性生成）。"""
        from scipy.interpolate import UnivariateSpline
        overlay = np.zeros((height, width, 3), dtype=np.uint8)
        lane_colors = {
            1: (0, 255,   0),
            2: (0, 200, 255),
            3: (255,  80, 200),
            4: (255, 200,   0),
            5: (80,  200, 255),
        }
        for lid, pts in sorted(lanes.items()):
            if len(pts) < 2:
                continue
            arr = np.array(pts, dtype=np.float64)
            ys, xs = arr[:, 1], arr[:, 0]
            order = np.argsort(ys)
            ys_u, xs_u = ys[order], xs[order]
            _, uid = np.unique(ys_u, return_index=True)
            ys_u, xs_u = ys_u[uid], xs_u[uid]
            try:
                k = min(3, len(ys_u) - 1)
                sp = UnivariateSpline(ys_u, xs_u, k=k, s=200 * len(ys_u))
                y_lo, y_hi = int(ys_u[0]), int(ys_u[-1])
                ys_e = np.arange(y_lo, y_hi + 1)
                xs_e = np.clip(sp(ys_e), 0, width - 1).astype(np.int32)
                curve = np.stack([xs_e, ys_e], axis=1).reshape(-1, 1, 2)
                color = lane_colors.get(lid, (200, 200, 200))
                cv2.polylines(overlay, [curve], False, color, 4)
            except Exception:
                pass
        return overlay

    @staticmethod
    def _assign_lane(
        cx: float, cy: float,
        lanes: dict[int, list[tuple[int, int]]],
    ) -> int | None:
        """根据 bbox 底部中心点判断车辆所属车道编号（区间序号，1-based）。

        返回 None 表示车辆不在任何标注车道内（对向车道或越界），不应计入统计。
        """
        from scipy.interpolate import UnivariateSpline
        import numpy as _np
        xs_at_cy: list[tuple[int, float]] = []
        for lid, pts in sorted(lanes.items()):
            if len(pts) < 2:
                continue
            arr = np.array(pts, dtype=np.float64)
            ys, xs = arr[:, 1], arr[:, 0]
            order = np.argsort(ys)
            ys_u, xs_u = ys[order], xs[order]
            _, uid = np.unique(ys_u, return_index=True)
            ys_u, xs_u = ys_u[uid], xs_u[uid]
            # cy 超出该线的 y 范围时跳过（而非直接返回 None）
            if cy < float(ys_u[0]) or cy > float(ys_u[-1]):
                continue
            try:
                k = min(3, len(ys_u) - 1)  # 点数不足4时降阶：2点→k=1线性
                sp = UnivariateSpline(ys_u, xs_u, k=k, s=200 * len(ys_u))
                xs_at_cy.append((lid, float(sp(cy))))
            except Exception:
                continue
        if len(xs_at_cy) < 2:
            return None
        xs_at_cy.sort(key=lambda t: t[1])
        for i in range(len(xs_at_cy) - 1):
            if xs_at_cy[i][1] <= cx <= xs_at_cy[i + 1][1]:
                return i + 1
        return None

    @staticmethod
    def _rewrite_dict_csv(path: Path, fields: list[str], rows: list[dict]) -> None:
        """用内存中的规范记录重写 CSV，避免并发/残留行污染最终文件。"""
        with Path(path).open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            for row in rows:
                writer.writerow({field: row.get(field, "") for field in fields})

    @staticmethod
    def _backfill_plates(
        traj_records: list[dict],
        cross_events: list[dict],
        plate_cache: dict[int, tuple[str, tuple[float, float, float, float] | None]],
    ) -> int:
        """按 track_id 将后续识别到的车牌回填到轨迹和断面事件。"""
        plate_by_tid = {
            int(tid): plate
            for tid, (plate, _) in plate_cache.items()
            if plate
        }
        filled = 0
        for rows in (traj_records, cross_events):
            for row in rows:
                try:
                    tid = int(row.get("track_id"))
                except (TypeError, ValueError):
                    continue
                plate = plate_by_tid.get(tid)
                if plate and not row.get("plate"):
                    row["plate"] = plate
                    filled += 1
        return filled

    def run(
        self,
        video_path: str | Path = _TEST_VIDEO,
        output_video: str | Path = _OUTPUT_VID,
        csv_path: str | Path = TRAJ_CSV_PATH,
        start_frame: int = _START_FRAME,
        end_frame: Optional[int] = _END_FRAME,
        sample_fps: int = TRAJ_SAMPLE_FPS,
        live_publisher=None,
        stop_event=None,
    ) -> Path:
        """运行跟踪并输出轨迹 CSV 和标注视频。

        Args:
            video_path:   输入视频路径。
            output_video: 输出带标注视频路径。
            csv_path:     输出轨迹 CSV 路径。
            start_frame:  起始帧号（含），默认 0。
            end_frame:    终止帧号（不含），None 表示处理到视频结尾。
            sample_fps:   每秒采样次数（轨迹记录频率）。
            live_publisher: 可选实时发布器，接收标注帧、进度和统计快照。
            stop_event:     可选停止信号，用于 Web 控制台请求优雅退出。

        Returns:
            写出的 CSV 文件路径。
        """
        _output_lock = _TrackerOutputLock(OUTPUT_DIR / ".tracker.lock").__enter__()
        cap  = open_video(video_path)
        meta = video_meta(cap)

        video_fps       = meta["fps"] if meta["fps"] > 0 else 25.0
        sample_interval = max(1, round(video_fps / sample_fps))   # 采样间隔（帧数）

        # 计算实际处理范围
        total_video = meta["frame_count"]
        start_frame = max(0, min(start_frame, total_video - 1))
        end_frame   = total_video if end_frame is None else min(end_frame, total_video)
        n_frames    = max(0, end_frame - start_frame)

        # Seek 到起始帧
        if start_frame > 0:
            cap.set(cv2.CAP_PROP_POS_FRAMES, start_frame)

        writer = AsyncWriter(make_writer(output_video, meta["fps"], 1920, 1080))

        print(f"[Tracker] 视频: {meta['width']}x{meta['height']}  {video_fps:.1f}fps  "
              f"处理帧: {start_frame} ~ {end_frame}（共{n_frames}帧）  "
              f"采样间隔: {sample_interval}帧")

        csv_path = Path(csv_path)
        csv_path.parent.mkdir(parents=True, exist_ok=True)

        class_ids    = list(VEHICLE_CLASSES.keys())
        fps_list: list[float] = []
        rows_written = 0
        _t_infer = _t_post = _t_cross = _t_draw = 0.0  # 各阶段累计耗时

        # 内存缓存，供最终 Excel 导出
        _all_traj_records: list[dict] = []
        _all_cross_events: list[dict] = []

        # ── 标定加载 + 测速器 + 车道线 overlay 初始化 ───────────────────────
        cal = get_calibration(video_path)
        if cal is None:
            print("[标定] ⚠ 加载失败，退化到默认值")
            _H = HOMOGRAPHY_MATRIX
            _lanes: dict[int, list] = {}
            _lane_types: dict[str, str] = {}
            _lane_overlay: np.ndarray | None = None
        else:
            _H = cal.homography
            _lanes = cal.lanes
            _lane_types: dict[str, str] = cal.metadata.get("lane_types", {})
            print(f"[标定] ✓ {cal.entrance} | 车道线 {len(_lanes)} 条 | "
                  f"H={cal.homography_method} | 光照={cal.lighting_preset}")
            _lane_overlay = self._build_lane_overlay(
                meta["height"], meta["width"], _lanes
            )

        # 测速器
        speed_est = SpeedEstimator(
            homography=_H,
            fps=video_fps,
            window=SPEED_WINDOW_FRAMES,
            pixels_per_meter=PIXELS_PER_METER,
            min_dist_m=SPEED_MIN_DIST_M,
        )

        # 轨迹分组器
        _entrance = cal.entrance if cal is not None else "未知"
        _grouper = TrajGrouper(
            interval_s=TRAJ_GROUP_INTERVAL_S,
            csv_path=TRAJ_GROUP_CSV_PATH,
            excel_path=EXCEL_REPORT_PATH,
            frame_width=meta["width"],
            frame_height=meta["height"],
        )

        # 断面检测器（复用同一 SpeedEstimator）
        _section_lines = _resolve_section_lines(video_path)
        section_det = CrossSectionDetector(
            _section_lines, _H, PIXELS_PER_METER, video_fps,
            speed_estimator=speed_est,
        )
        _section_hit: dict[str, int] = {}
        _HIGHLIGHT_FRAMES = 8

        cross_path = Path(CROSS_SECTION_CSV_PATH)
        cross_path.parent.mkdir(parents=True, exist_ok=True)
        cross_f = cross_path.open("w", newline="", encoding="utf-8")
        cross_writer = csv.DictWriter(cross_f, fieldnames=self._CROSS_CSV_FIELDS)
        cross_writer.writeheader()

        stats_path = Path(VEHICLE_STATS_CSV_PATH)
        stats_path.parent.mkdir(parents=True, exist_ok=True)
        stats_fh = open(stats_path, 'w', newline='', encoding='utf-8')
        stats_writer = csv.writer(stats_fh)
        stats_writer.writerow([
            "track_id", "first_frame", "last_frame", "lane_id",
            "avg_speed_kmh", "max_speed_kmh", "min_speed_kmh", "n_samples",
        ])

        last_known_lane: dict[int, int | None] = {}
        active_tids: set[int] = set()
        prev_active: set[int] = set()
        _traj_buf: dict[int, list[tuple[float, float]]] = {}
        _tid_cls: dict[int, str] = {}
        _grace_buf: dict[int, int] = {}   # tid → 剩余存活帧数
        _live_vehicle_ids: set[int] = set()
        _live_recent_events: list[dict] = []
        _LIVE_FRAME_INTERVAL = max(1, round(video_fps / _LIVE_PREVIEW_FPS))

        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer_csv = csv.DictWriter(f, fieldnames=self._CSV_FIELDS)
            writer_csv.writeheader()

            for local_idx, frame in iter_frames(cap, n_frames):
                if stop_event is not None and stop_event.is_set():
                    print("[Tracker] 收到停止请求，正在结束检测")
                    break

                frame_idx = start_frame + local_idx   # 视频中的绝对帧号
                t0 = time.perf_counter()

                # ByteTrack 跟踪推理（persist=True 保持跨帧状态）
                results = self._model.track(
                    frame,
                    classes=class_ids,
                    device=self.device,
                    conf=self.conf,
                    persist=True,
                    tracker="botsort.yaml",
                    verbose=False,
                    half=True,
                )[0]
                t1 = time.perf_counter()
                _t_infer += t1 - t0

                cur_fps = 1.0 / (t1 - t0)
                fps_list.append(cur_fps)

                timestamp_s   = round(frame_idx / video_fps, 3)
                should_sample = (frame_idx % sample_interval == 0)

                boxes_xyxy, labels, confs, track_ids, plates, plate_boxes = [], [], [], [], [], []
                _live_class_counts: dict[str, int] = {}
                _live_lane_counts: dict[str, int] = {}
                _live_speed_samples: list[float] = []

                boxes_data = results.boxes
                if boxes_data is not None and len(boxes_data):
                    ids = (boxes_data.id.int().tolist()
                           if boxes_data.id is not None
                           else [None] * len(boxes_data))

                    for i, box in enumerate(boxes_data):
                        cls_id = int(box.cls[0])
                        if cls_id not in VEHICLE_CLASSES:
                            continue
                        label = VEHICLE_CLASSES[cls_id]
                        conf  = float(box.conf[0])
                        tid   = ids[i]
                        x1, y1, x2, y2 = map(int, box.xyxy[0])

                        boxes_xyxy.append((x1, y1, x2, y2))
                        labels.append(label)
                        confs.append(conf)
                        track_ids.append(tid)
                        _live_class_counts[label] = _live_class_counts.get(label, 0) + 1

                        # 每帧：对未锁定车辆积累投票，锁定后不再调用 HyperLPR3
                        if (tid is not None
                                and label in _PLATE_VEHICLE_CLASSES
                                and tid not in self._plate_rec._cache):
                            self._plate_rec.recognize(frame, tid, (x1, y1, x2, y2))

                        # 每帧：从缓存取已确认车牌，用当前 bbox 还原绝对坐标
                        cached = (
                            self._plate_rec._cache.get(tid, ("", None))
                            if tid is not None else ("", None)
                        )
                        plate_display, cached_rel_box = cached
                        plate_box_display = (
                            _rel_to_abs(cached_rel_box, x1, y1, x2, y2)
                            if cached_rel_box is not None else None
                        )

                        # 测速 + 车道归属（必须在轨迹采样前完成，保证本帧速度/车道已更新）
                        cx_c = float((x1 + x2) / 2)
                        cy_b = float(y2)
                        v_now: float | None = None
                        lane_id: int | None = None
                        if tid is not None:
                            speed_est.update(frame_idx, tid, cx_c, cy_b)
                            v_now = speed_est.instant_speed(tid)
                            lane_id = self._assign_lane(cx_c, cy_b, _lanes) if _lanes else None
                            if lane_id is not None:
                                last_known_lane[tid] = lane_id
                                _live_lane_counts[str(lane_id)] = _live_lane_counts.get(str(lane_id), 0) + 1
                            active_tids.add(tid)
                            _live_vehicle_ids.add(tid)
                            if tid in _grace_buf:
                                del _grace_buf[tid]   # 从 grace 复活
                            _traj_buf.setdefault(tid, []).append((cx_c, float((y1 + y2) / 2)))
                            _tid_cls[tid] = label
                            if v_now is not None and v_now > 0:
                                _live_speed_samples.append(float(v_now))

                            # 速度+车道标签（bbox 下方，避免与类别标签重叠）
                            parts = []
                            if lane_id is not None:
                                parts.append(f"L{lane_id}")
                            if v_now is not None and v_now > 0:
                                parts.append(f"{v_now:.0f}km/h")
                            if parts:
                                label_speed = " ".join(parts)
                                put_text(frame, label_speed,
                                         (int(x1), min(frame.shape[0] - 20, int(y2) + 24)),
                                         color=(0, 255, 255),
                                         font_scale=0.75)

                        # 轨迹采样记录（speed/lane 已是本帧最新值）
                        # lane_id 为空表示对向/越界车辆，仍记录轨迹，但统计时会过滤
                        _cur_lane = last_known_lane.get(tid)
                        if should_sample and tid is not None:
                            cx = int(cx_c)
                            cy = (y1 + y2) // 2
                            plate_display, rel_box = self._plate_rec._cache.get(tid, ("", None))
                            plate_box_display = (
                                _rel_to_abs(rel_box, x1, y1, x2, y2)
                                if rel_box is not None else None
                            )
                            _speed_kmh = round(v_now, 1) if v_now is not None else None
                            _lane_type = _lane_types.get(str(_cur_lane), "") if _cur_lane else ""
                            _row = {
                                "frame_id":    frame_idx,
                                "timestamp_s": timestamp_s,
                                "track_id":    tid,
                                "class_name":  label,
                                "lane_id":     _cur_lane,
                                "lane_type":   _lane_type,
                                "cx":          cx,
                                "cy":          cy,
                                "x1":          x1,
                                "y1":          y1,
                                "x2":          x2,
                                "y2":          y2,
                                "speed_kmh":   _speed_kmh,
                                "plate":       plate_display,
                            }
                            writer_csv.writerow(_row)
                            _all_traj_records.append(_row)
                            rows_written += 1

                        plates.append(plate_display)
                        plate_boxes.append(plate_box_display)

                t2 = time.perf_counter()
                _t_post += t2 - t1

                # ── 断面过车检测（Position B）──────────────────────────────────
                for _box, _label, _tid in zip(boxes_xyxy, labels, track_ids):
                    if _tid is None:
                        continue
                    _bx1, _by1, _bx2, _by2 = _box
                    for ev in section_det.update(
                        frame_idx, timestamp_s, _tid, _label,
                        frame, _bx1, _by1, _bx2, _by2,
                    ):
                        # 将已识别车牌绑定到过线事件（按 track_id 查缓存）
                        ev["plate"] = self._plate_rec._cache.get(_tid, ("", None))[0]
                        cross_writer.writerow(ev)
                        _all_cross_events.append(ev)
                        _live_recent_events.append({
                            "timestamp_s": ev.get("timestamp_s", timestamp_s),
                            "section": ev.get("section", ""),
                            "direction": ev.get("direction", ""),
                            "class_name": ev.get("class_name", ""),
                            "speed_kmh": ev.get("speed_kmh", 0),
                            "plate": ev.get("plate", ""),
                        })
                        _live_recent_events = _live_recent_events[-8:]
                        _section_hit[ev["section"]] = frame_idx

                t3 = time.perf_counter()
                _t_cross += t3 - t2

                # ── 缩放至 1080p 再绘制（4K 画图是最大瓶颈）─────────────────
                S = 0.5  # 3840→1920, 2160→1080
                frame = cv2.resize(frame, (1920, 1080))
                _lane_1080 = cv2.resize(_lane_overlay, (1920, 1080)) if _lane_overlay is not None else None
                _sec_lines_1080 = [(_n, int(_a*S), int(_b*S), int(_c*S), int(_d*S), _e, _f)
                                   for _n, _a, _b, _c, _d, _e, _f in _section_lines]
                _boxes_1080 = [(int(x1*S), int(y1*S), int(x2*S), int(y2*S)) for (x1, y1, x2, y2) in boxes_xyxy]
                _plates_1080 = [(int(p[0]*S), int(p[1]*S), int(p[2]*S), int(p[3]*S)) if p else None
                                for p in plate_boxes]

                # ── 绘制断面线（在车辆框下层）────────────────────────────────
                for _name, _lx1, _ly1, _lx2, _ly2, _, _ in _sec_lines_1080:
                    _age = frame_idx - _section_hit.get(_name, -999)
                    _col = (0, 255, 255) if _age <= _HIGHLIGHT_FRAMES else (0, 0, 255)
                    _thi = 2              if _age <= _HIGHLIGHT_FRAMES else 1
                    cv2.line(frame, (_lx1, _ly1), (_lx2, _ly2), _col, _thi)
                    put_text(frame, _name, (_lx1, _ly1 - 12), _col)

                # 画真实车牌框和识别标签（含中文省份字符，走 PIL 渲染）
                for plate_str, plate_box in zip(plates, _plates_1080):
                    if not plate_str or plate_box is None:
                        continue
                    px1, py1, px2, py2 = plate_box
                    cv2.rectangle(frame, (px1, py1), (px2, py2), (0, 255, 255), 1)
                    put_text(frame, plate_str, (px1, max(0, py1 - 11)), (0, 255, 255))

                # 绘制带跟踪ID和车牌的标注框
                draw_boxes(frame, _boxes_1080, labels, confs, track_ids, plates)
                put_fps_text(frame, cur_fps, len(_boxes_1080))

                # 车道线 overlay 叠加
                annotated = frame
                if _lane_1080 is not None:
                    annotated = cv2.addWeighted(frame, 1.0, _lane_1080, 0.5, 0)

                if live_publisher is not None:
                    avg_live_speed = (
                        sum(_live_speed_samples) / len(_live_speed_samples)
                        if _live_speed_samples else 0.0
                    )
                    live_publisher.update_progress(frame_idx, timestamp_s, cur_fps)
                    live_publisher.publish_stats({
                        "vehicles": len(_live_vehicle_ids),
                        "events": len(_all_cross_events),
                        "avg_speed": round(avg_live_speed, 1),
                        "active_tracks": sum(1 for tid in track_ids if tid is not None),
                        "recent_events": list(reversed(_live_recent_events)),
                        "class_counts": _live_class_counts,
                        "lane_counts": _live_lane_counts,
                    })
                    if frame_idx % _LIVE_FRAME_INTERVAL == 0:
                        live_publisher.publish_frame(annotated, width=960, quality=68)

                # 消失的 track → 进入 grace buffer
                for tid in (prev_active - active_tids):
                    if tid not in _grace_buf:
                        _grace_buf[tid] = _GRACE_FRAMES

                # grace 到期 → 真正结束
                expired = [tid for tid, c in _grace_buf.items() if c <= 0]
                for tid in expired:
                    del _grace_buf[tid]
                    s = speed_est.finalize(tid)
                    if s and s['n_samples'] >= SPEED_MIN_SAMPLES:
                        stats_writer.writerow([
                            tid, s['first_frame'], s['last_frame'],
                            last_known_lane.get(tid),
                            round(s['avg_kmh'], 1), round(s['max_kmh'], 1),
                            round(s['min_kmh'], 1), s['n_samples'],
                        ])
                    _pts = _traj_buf.pop(tid, [])
                    _lid = last_known_lane.get(tid)
                    _grouper.on_track_end(
                        tid, _pts, _tid_cls.pop(tid, "car"),
                        _lid,
                        _lane_types.get(str(_lid), ""),
                        _entrance,
                    )
                    last_known_lane.pop(tid, None)

                # 递减未到期的 grace 计数
                for tid in list(_grace_buf):
                    _grace_buf[tid] -= 1

                prev_active = set(active_tids)
                active_tids.clear()

                writer.write(annotated)

                t4 = time.perf_counter()
                _t_draw += t4 - t3

                _grouper.tick(timestamp_s)

                if (local_idx + 1) % 30 == 0:
                    avg30 = sum(fps_list[-30:]) / min(len(fps_list), 30)
                    pct   = (local_idx + 1) / n_frames * 100
                    n = min(local_idx + 1, 30)
                    print(f"[{pct:5.1f}%] 帧 {frame_idx:4d}/{end_frame-1}  "
                          f"FPS:{avg30:.1f} "
                          f"推理:{_t_infer/n*1000:.0f}ms "
                          f"后处理:{_t_post/n*1000:.0f}ms "
                          f"断面:{_t_cross/n*1000:.0f}ms "
                          f"绘制:{_t_draw/n*1000:.0f}ms "
                          f"轨迹:{rows_written}", flush=True)
                    _t_infer = _t_post = _t_cross = _t_draw = 0.0

        # finalize grace buffer 中的 track（视频结束，强制到期）
        for tid in list(_grace_buf):
            del _grace_buf[tid]
            s = speed_est.finalize(tid)
            if s and s['n_samples'] >= SPEED_MIN_SAMPLES:
                stats_writer.writerow([
                    tid, s['first_frame'], s['last_frame'],
                    last_known_lane.get(tid),
                    round(s['avg_kmh'], 1), round(s['max_kmh'], 1),
                    round(s['min_kmh'], 1), s['n_samples'],
                ])
            _pts = _traj_buf.pop(tid, [])
            _lid = last_known_lane.get(tid)
            _grouper.on_track_end(
                tid, _pts, _tid_cls.pop(tid, "car"),
                _lid,
                _lane_types.get(str(_lid), ""),
                _entrance,
            )
            last_known_lane.pop(tid, None)

        # finalize 剩余 track
        for tid in list(speed_est._tracks.keys()):
            s = speed_est.finalize(tid)
            if s and s['n_samples'] >= SPEED_MIN_SAMPLES:
                stats_writer.writerow([
                    tid, s['first_frame'], s['last_frame'],
                    last_known_lane.get(tid),
                    round(s['avg_kmh'], 1), round(s['max_kmh'], 1),
                    round(s['min_kmh'], 1), s['n_samples'],
                ])
        stats_fh.close()
        print(f"[Stats] vehicle_stats.csv: {stats_path}")

        cap.release()
        writer.release()
        cross_f.close()

        avg_fps = sum(fps_list) / len(fps_list) if fps_list else 0.0
        print(f"\n=== 轨迹提取完成 ===")
        print(f"处理帧数    : {len(fps_list)}")
        print(f"平均推理FPS : {avg_fps:.1f}")
        print(f"轨迹记录行  : {rows_written}")
        print(f"CSV 输出    : {csv_path}")
        print(f"断面CSV     : {cross_path}")
        print(f"视频输出    : {output_video}")

        # ── 断面过车摘要 ─────────────────────────────────────────────────────
        from collections import defaultdict as _dd
        _sec_cat: dict[str, dict[str, int]] = _dd(lambda: _dd(int))
        for _ev in _all_cross_events:
            _sec_cat[_ev["section"]][_ev.get("vehicle_category", "未知")] += 1
        unique_tids = len({r["track_id"] for r in _all_traj_records})
        total_motor    = sum(e.get("vehicle_category") == "机动车"  for e in _all_cross_events)
        total_nonmotor = sum(e.get("vehicle_category") == "非机动车" for e in _all_cross_events)

        print(f"\n=== 断面过车摘要 ===")
        for sec, cats in sorted(_sec_cat.items()):
            motor    = cats.get("机动车",  0)
            nonmotor = cats.get("非机动车", 0)
            print(f"  {sec:<18} 机动车 {motor:3d}辆   非机动车 {nonmotor:3d}辆   合计 {motor+nonmotor:3d}辆")
        print(f"跟踪车辆总数    : {unique_tids} 辆（唯一 track_id）")
        print(f"断面过车合计    : 机动车 {total_motor} 辆  /  非机动车 {total_nonmotor} 辆")

        filled_plates = self._backfill_plates(
            _all_traj_records,
            _all_cross_events,
            self._plate_rec._cache,
        )
        self._rewrite_dict_csv(csv_path, self._CSV_FIELDS, _all_traj_records)
        self._rewrite_dict_csv(cross_path, self._CROSS_CSV_FIELDS, _all_cross_events)
        if filled_plates:
            print(f"[Plate] 已按 track_id 回填车牌 {filled_plates} 处，并重写 CSV")

        self._export_excel(
            _all_traj_records, _all_cross_events,
            self._CSV_FIELDS, self._CROSS_CSV_FIELDS,
        )
        _grouper.finalize()
        _output_lock.__exit__(None, None, None)
        return csv_path

    def _export_excel(
        self,
        traj_records: list[dict],
        cross_events: list[dict],
        traj_fields: list[str],
        cross_fields: list[str],
        output_path: Path = EXCEL_REPORT_PATH,
    ) -> None:
        """将所有数据导出为多 Sheet Excel 报表。"""
        try:
            import openpyxl
        except ImportError:
            print("[Excel] 未找到 openpyxl，跳过 Excel 导出（pip install openpyxl）")
            return

        from collections import defaultdict

        _CN = {
            "frame_id":          "帧编号",
            "timestamp_s":       "时间戳(秒)",
            "section":           "断面名称",
            "arrival_departure": "到达/离去",
            "track_id":          "车辆ID",
            "class_name":        "车辆类型",
            "vehicle_category":  "车辆性质",
            "lane_id":           "车道编号",
            "color":             "车身颜色",
            "direction":         "行驶方向",
            "speed_kmh":         "速度(km/h)",
            "headway_s":         "车头时距(秒)",
            "spacing_m":         "车头间距(米)",
            "cx":                "中心X",
            "cy":                "中心Y",
            "x1":                "左上X",
            "y1":                "左上Y",
            "x2":                "右下X",
            "y2":                "右下Y",
            "plate":             "车牌号",
        }

        def _headers(fields: list[str]) -> list[str]:
            return [_CN.get(f, f) for f in fields]

        def _valid_lane_id(value) -> bool:
            """Excel 每车道统计只应使用已归属到标注车道的轨迹点。"""
            if value is None or value == "":
                return False
            try:
                return not np.isnan(value)
            except TypeError:
                return True

        def _bbox_iou(a: dict, b: dict) -> float:
            ax1, ay1, ax2, ay2 = (float(a.get(k, 0) or 0) for k in ("x1", "y1", "x2", "y2"))
            bx1, by1, bx2, by2 = (float(b.get(k, 0) or 0) for k in ("x1", "y1", "x2", "y2"))
            ix1, iy1 = max(ax1, bx1), max(ay1, by1)
            ix2, iy2 = min(ax2, bx2), min(ay2, by2)
            iw, ih = max(0.0, ix2 - ix1), max(0.0, iy2 - iy1)
            inter = iw * ih
            area_a = max(0.0, ax2 - ax1) * max(0.0, ay2 - ay1)
            area_b = max(0.0, bx2 - bx1) * max(0.0, by2 - by1)
            denom = area_a + area_b - inter
            return inter / denom if denom > 0 else 0.0

        def _center_dist(a: dict, b: dict) -> float:
            return float(np.hypot(
                float(a.get("cx", 0) or 0) - float(b.get("cx", 0) or 0),
                float(a.get("cy", 0) or 0) - float(b.get("cy", 0) or 0),
            ))

        def _dedupe_snapshot(records: list[dict]) -> list[dict]:
            """去掉同一快照内由 ID 跳变/类别抖动造成的重复车辆。"""
            kept: list[dict] = []
            for rec in sorted(records, key=lambda r: VEHICLE_LENGTHS_M.get(r.get("class_name", "car"), 4.5)):
                if any(_bbox_iou(rec, old) >= 0.85 or _center_dist(rec, old) <= 5.0 for old in kept):
                    continue
                kept.append(rec)
            return kept

        wb = openpyxl.Workbook()

        # ── Sheet1: 断面过车 ─────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "断面过车"
        ws1.append(_headers(cross_fields))
        for ev in cross_events:
            ws1.append([ev.get(f, "") for f in cross_fields])

        # ── Sheet2: 车辆轨迹 ─────────────────────────────────────────────────
        ws2 = wb.create_sheet("车辆轨迹")
        ws2.append(_headers(traj_fields))
        for rec in traj_records:
            ws2.append([rec.get(f, "") for f in traj_fields])

        # ── Sheet3: 流量统计 Q = N/T（每分钟 × 断面 × 车辆性质）────────────
        ws3 = wb.create_sheet("流量统计")
        ws3.append(["统计起始(秒)", "统计结束(秒)", "断面名称", "车辆性质", "过车数", "流量(辆/分钟)"])
        flow_count: dict[tuple, int] = defaultdict(int)
        for ev in cross_events:
            win = int(float(ev.get("timestamp_s", 0)) // 60)
            flow_count[(win, ev.get("section", ""), ev.get("vehicle_category", ""))] += 1
        for (win, section, cat), count in sorted(flow_count.items()):
            ws3.append([win * 60, (win + 1) * 60, section, cat, count, round(count / 1.0, 2)])

        # ── Sheet4: 空间占有率（每5秒 × 每车道）────────────────────────────
        ws4 = wb.create_sheet("空间占有率")
        ws4.append([
            "统计起始(秒)", "统计结束(秒)", "车道编号",
            "平均在场车辆数", "平均占用总长度(米)", "平均空间占有率(%)",
        ])
        # 先按每秒快照统计，再对 5 秒窗口求平均；不能把 5 秒内出现过的 track_id 做并集。
        occ_snapshot: dict[tuple, list] = defaultdict(list)
        for rec in traj_records:
            lane = rec.get("lane_id", "")
            if not _valid_lane_id(lane):
                continue
            ts = int(float(rec.get("timestamp_s", 0)))
            occ_snapshot[(ts, lane)].append(rec)

        occ_window: dict[tuple, list] = defaultdict(list)
        for (ts, lane), records in occ_snapshot.items():
            deduped = _dedupe_snapshot(records)
            total_len = sum(VEHICLE_LENGTHS_M.get(r.get("class_name", "car"), 4.5) for r in deduped)
            occ_pct = round(min(100.0, total_len / SECTION_ROAD_LENGTH_M * 100), 1)
            occ_window[(ts // 5, lane)].append((len(deduped), total_len, occ_pct))

        for (win, lane) in sorted(occ_window.keys(), key=lambda t: (t[0], str(t[1]))):
            samples = occ_window[(win, lane)]
            n_samples = len(samples)
            avg_count = round(sum(v[0] for v in samples) / n_samples, 1)
            avg_len = round(sum(v[1] for v in samples) / n_samples, 1)
            avg_occ = round(sum(v[2] for v in samples) / n_samples, 1)
            ws4.append([win * 5, (win + 1) * 5, lane, avg_count, avg_len, avg_occ])

        # ── Sheet5: 排队长度（每秒 × 机动车道，仅机动车）
        # 排队判定：speed ≤ QUEUE_SPEED_THRESH_KMH
        # 排队长度(米) = Σ各车车身长度 + (n-1) × QUEUE_GAP_M
        _MOTOR_CLASSES = {"car", "bus", "truck"}
        ws5 = wb.create_sheet("排队长度")
        ws5.append([
            "统计起始(秒)", "统计结束(秒)", "车道编号",
            "排队车辆数", "排队长度(米)", "平均排队车速(km/h)",
        ])
        # (秒, lane_id) -> list[rec]；同一秒内先做空间去重，避免 ID 跳变重复计数。
        queue_bucket: dict[tuple, list] = defaultdict(list)
        for rec in traj_records:
            if rec.get("class_name") not in _MOTOR_CLASSES:
                continue
            if rec.get("lane_type") != "motor":
                continue
            spd_raw = rec.get("speed_kmh")
            if spd_raw is None or spd_raw == "":
                continue
            spd = float(spd_raw)
            if 0.0 <= spd <= QUEUE_SPEED_THRESH_KMH:
                lane = rec.get("lane_id", "")
                if not _valid_lane_id(lane):
                    continue
                ts = int(float(rec.get("timestamp_s", 0)))
                queue_bucket[(ts, lane)].append(rec)

        for (ts, lane) in sorted(queue_bucket.keys(), key=lambda t: (t[0], str(t[1]))):
            records = _dedupe_snapshot(queue_bucket[(ts, lane)])
            n = len(records)
            avg_spd = round(sum(float(r.get("speed_kmh", 0) or 0) for r in records) / n, 1)
            total_body = sum(VEHICLE_LENGTHS_M.get(r.get("class_name", "car"), 4.5) for r in records)
            queue_len  = round(total_body + max(0, n - 1) * QUEUE_GAP_M, 1)
            ws5.append([
                ts, ts + 1,
                lane,
                n, queue_len, avg_spd,
            ])

        wb.save(output_path)
        print(f"[Excel] 报表已导出: {output_path}")


if __name__ == "__main__":
    tracker = TrajectoryTracker()
    tracker.run(video_path=_TEST_VIDEO, start_frame=_START_FRAME, end_frame=_END_FRAME)
